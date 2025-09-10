#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
연구비 처리 자동화 프로그램 - 핵심 데이터 처리 로직

작성자: 차세대지원팀 데이터 김훈민
작성일자: 2025-07-22
"""

import os
import pandas as pd
import logging
from typing import Dict, Optional, List, Tuple
from openpyxl.chart import BarChart, PieChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, IconSetRule
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import (
    BUSINESS_PREFIX, RESEARCH_PREFIX, SUPPORTED_EXTENSIONS,
    SUMMARY_COLUMN, UNCLASSIFIED_WARNING_THRESHOLD,
    OUTPUT_SHEET_NAMES, OUTPUT_COLUMNS, RESEARCH_ADDITIONAL_COLUMNS,
    EXCEL_STYLING, BUDGET_CLASSIFICATION, SUMMARY_SHEET_COLUMNS, TOTAL_SHEET_COLUMNS
)


class ExcelFileLoader:
    '''Excel 파일 로더 클래스'''

    def __init__(self):
        self.file_path = None
        self.data = None

    def load_file(self, file_path: str) -> bool:
        '''Excel 파일을 로드합니다.'''
        try:
            if not self._validate_file(file_path):
                return False

            self.file_path = file_path
            self.data = pd.read_excel(file_path)

            if self.data.empty:
                logging.warning("로드된 데이터가 비어있습니다.")
                return False

            logging.info(f"파일 로드 성공: {file_path}")
            logging.info(f"데이터 크기: {self.data.shape}")
            return True

        except (pd.errors.EmptyDataError, pd.errors.ParserError) as e:
            logging.error(f"Excel 파일 처리 오류: {str(e)}")
            return False
        except PermissionError:
            logging.error("파일 접근 권한이 없습니다.")
            return False
        except Exception as e:
            logging.error(f"파일 로드 실패: {str(e)}")
            return False

    def _validate_file(self, file_path: str) -> bool:
        '''파일 유효성 검증'''
        if not os.path.exists(file_path):
            logging.error(f"파일이 존재하지 않습니다: {file_path}")
            return False

        if not any(file_path.lower().endswith(ext) for ext in SUPPORTED_EXTENSIONS):
            logging.error(f"지원하지 않는 파일 형식입니다: {file_path}")
            return False

        return True

    def get_data_info(self) -> Optional[Dict]:
        '''로드된 데이터의 정보를 반환합니다.'''
        if self.data is None:
            return None

        return {
            'file_path': self.file_path,
            'shape': self.data.shape,
            'columns': list(self.data.columns),
            'sample_data': self.data.head().to_dict('records')
        }


class DataClassifier:
    '''데이터 분류 클래스'''

    def __init__(self):
        self.business_data = None
        self.research_data = None
        self.unclassified_data = None
        self.classification_stats = {}

    def classify_data(self, data: pd.DataFrame, 
                     summary_column: str = SUMMARY_COLUMN) -> Dict[str, pd.DataFrame]:
        '''데이터를 사업비와 연구비로 분류합니다.'''
        if data is None or data.empty:
            raise ValueError("분류할 데이터가 없습니다.")

        if summary_column not in data.columns:
            available_columns = ', '.join(data.columns.tolist())
            raise ValueError(
                f"'{summary_column}' 컬럼이 데이터에 존재하지 않습니다.\n"
                f"사용 가능한 컬럼: {available_columns}"
            )

        logging.info(f"데이터 분류 시작 - 총 {len(data)}건")

        # 데이터 전처리
        work_data = data.copy()
        work_data[summary_column] = work_data[summary_column].fillna('').astype(str)

        # 분류 수행
        business_mask = work_data[summary_column].str.startswith(BUSINESS_PREFIX) # 25 차세대
        research_mask = work_data[summary_column].str.startswith(RESEARCH_PREFIX) # 25 심층연구

        self.business_data = work_data[business_mask].copy()
        self.research_data = work_data[research_mask].copy()
        self.unclassified_data = work_data[~(business_mask | research_mask)].copy() # 미분류는 대부분 인건비 일 것이라고 추정.

        # 통계 생성
        self._generate_stats(len(data))
        self._log_results()

        return {
            'business': self.business_data,
            'research': self.research_data,
            'unclassified': self.unclassified_data
        }

    def _generate_stats(self, total_count: int) -> None:
        '''분류 통계 정보를 생성합니다.'''
        business_count = len(self.business_data)
        research_count = len(self.research_data)
        unclassified_count = len(self.unclassified_data)

        self.classification_stats = {
            'total': total_count,
            'business_count': business_count,
            'research_count': research_count,
            'unclassified_count': unclassified_count,
            'business_percentage': (business_count / total_count * 100) if total_count > 0 else 0,
            'research_percentage': (research_count / total_count * 100) if total_count > 0 else 0,
            'unclassified_percentage': (unclassified_count / total_count * 100) if total_count > 0 else 0
        }

    def _log_results(self) -> None:
        '''분류 결과를 로깅합니다.'''
        stats = self.classification_stats
        logging.info(f"분류 완료 - 사업비: {stats['business_count']}건, "
                    f"연구비: {stats['research_count']}건, "
                    f"미분류: {stats['unclassified_count']}건")

        if stats['unclassified_count'] > stats['total'] * UNCLASSIFIED_WARNING_THRESHOLD:
            logging.warning(f"미분류 데이터가 {stats['unclassified_count']}건으로 많습니다.")

    def get_classification_stats(self) -> Dict:
        '''분류 통계 정보를 반환합니다.'''
        return self.classification_stats.copy()

    def get_business_data(self) -> Optional[pd.DataFrame]:
        '''사업비 데이터를 반환합니다.'''
        return self.business_data.copy() if self.business_data is not None else None

    def get_research_data(self) -> Optional[pd.DataFrame]:
        '''연구비 데이터를 반환합니다.'''
        return self.research_data.copy() if self.research_data is not None else None

    def get_unclassified_data(self) -> Optional[pd.DataFrame]:
        '''분류되지 않은 데이터를 반환합니다.'''
        return self.unclassified_data.copy() if self.unclassified_data is not None else None


class ExcelExporter:
    '''Excel 파일 출력 클래스'''

    def __init__(self):
        self.business_data = None
        self.research_data = None

    def export_to_excel(self, business_data: pd.DataFrame, research_data: pd.DataFrame,
                       output_path: str) -> bool:
        '''분류된 데이터를 Excel 파일로 출력합니다.'''
        try:
            self.business_data = business_data
            self.research_data = research_data

            # Excel writer 생성
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

                # 데이터 시트들 생성
                total_generator = TotalSheetGenerator()
                total_sheet = total_generator.generate_total_sheet(business_data, research_data)
                summary_generator = SummarySheetGenerator()
                dashboard_generator = DashboardGenerator()
                business_summary_sheet = pd.DataFrame()
                research_summary_sheet = pd.DataFrame()

                # 1. 대시보드 시트 생성 (첫 번째)
                if not total_sheet.empty:
                    dashboard_sheet = dashboard_generator.generate_dashboard_sheet(total_sheet)
                    dashboard_sheet.to_excel(
                        writer,
                        sheet_name='대시보드',
                        index=False
                    )
                    # 대시보드 워크시트에 실제 대시보드 생성
                    dashboard_generator.create_dashboard_in_worksheet(
                        writer.sheets['대시보드'],
                        total_sheet
                    )
                    
                # 2. 총액 시트 생성 (두 번째)
                if not total_sheet.empty:
                    total_sheet.to_excel(
                        writer,
                        sheet_name='총액',
                        index=False
                    )
                    # 총액 시트 스타일링 적용
                    self._apply_total_sheet_styling(
                        writer.sheets['총액'],
                        total_sheet
                    )
                    logging.info(f"총액 시트 생성 완료: {len(total_sheet)}건")

                # 2. 사업비 요약 시트 생성 (세 번째)
                if business_data is not None and not business_data.empty:
                    business_summary_sheet = summary_generator.generate_summary_sheet(business_data)
                    if not business_summary_sheet.empty:
                        business_summary_sheet.to_excel(
                            writer,
                            sheet_name='사업비',
                            index=False
                        )
                        # 사업비 요약 시트 스타일링 적용
                        self._apply_summary_sheet_styling(
                            writer.sheets['사업비'],
                            business_summary_sheet
                        )
                        logging.info(f"사업비 요약 시트 생성 완료: {len(business_summary_sheet)}건")

                # 3. 연구비 요약 시트 생성 (네 번째)
                if research_data is not None and not research_data.empty:
                    research_summary_sheet = summary_generator.generate_research_summary_sheet(research_data)
                    if not research_summary_sheet.empty:
                        research_summary_sheet.to_excel(
                            writer,
                            sheet_name='연구비',
                            index=False
                        )
                        # 연구비 요약 시트 스타일링 적용
                        self._apply_summary_sheet_styling(
                            writer.sheets['연구비'],
                            research_summary_sheet
                        )
                        logging.info(f"연구비 요약 시트 생성 완료: {len(research_summary_sheet)}건")

                # 4. 집행관리(사업비) 시트 생성 (다섯 번째)
                if business_data is not None and not business_data.empty:
                    business_sheet = self._prepare_business_sheet(business_data)
                    business_sheet.to_excel(
                        writer,
                        sheet_name=OUTPUT_SHEET_NAMES['business'],
                        index=False
                    )
                    # 사업비 시트 스타일링 적용
                    self._apply_sheet_styling(
                        writer.sheets[OUTPUT_SHEET_NAMES['business']],
                        business_sheet,
                        'business'
                    )
                    logging.info(f"집행관리(사업비) 시트 생성 완료: {len(business_sheet)}건")

                # 5. 집행관리(연구비) 시트 생성 (여섯 번째)
                if research_data is not None and not research_data.empty:
                    research_sheet = self._prepare_research_sheet(research_data)
                    research_sheet.to_excel(
                        writer,
                        sheet_name=OUTPUT_SHEET_NAMES['research'],
                        index=False
                    )
                    # 연구비 시트 스타일링 적용
                    self._apply_sheet_styling(
                        writer.sheets[OUTPUT_SHEET_NAMES['research']],
                        research_sheet,
                        'research'
                    )
                    logging.info(f"집행관리(연구비) 시트 생성 완료: {len(research_sheet)}건")

                # 6. 시트 순서 조정 - 대시보드를 첫 번째로 이동
                self._reorder_sheets_with_dashboard_first(writer.book)

            logging.info(f"Excel 파일 출력 완료: {output_path}")
            
            # 7. xlwings를 사용한 대화형 피벗 테이블 추가
            from config import ENABLE_INTERACTIVE_PIVOT
            if ENABLE_INTERACTIVE_PIVOT:
                pivot_generator = InteractivePivotGenerator()
                if pivot_generator.xlwings_available:
                    # 기존 피벗 테이블 추가
                    success = pivot_generator.add_interactive_features(output_path)
                    if success:
                        logging.info("대화형 피벗 테이블 추가 완료")
                    else:
                        logging.warning("대화형 피벗 테이블 추가 실패")
                    
                    # 연도별 예산 비교 피벗 테이블 추가
                    yearly_success = pivot_generator.create_yearly_budget_comparison(output_path)
                    if yearly_success:
                        logging.info("연도별 예산 비교 테이블 추가 완료")
                    else:
                        logging.warning("연도별 예산 비교 테이블 추가 실패")
                    
                    # 모든 시트 순서 최종 조정
                    reorder_success = pivot_generator.reorder_all_sheets(output_path)
                    if reorder_success:
                        logging.info("전체 시트 순서 조정 완료")
                    else:
                        logging.warning("전체 시트 순서 조정 실패")
                else:
                    logging.info("xlwings 사용 불가 - 정적 시트만 생성됨")
            else:
                logging.info("대화형 피벗 테이블 기능이 비활성화됨")
                
            return True

        except Exception as e:
            logging.error(f"Excel 파일 출력 실패: {str(e)}")
            return False

    def _reorder_sheets_with_dashboard_first(self, workbook):
        '''시트 순서를 원하는 순서로 조정합니다.'''
        try:
            # 원하는 시트 순서 정의
            desired_order = [
                '대시보드',
                '총액',
                '사업비',
                '연구비',
                '집행관리(사업비)',
                '집행관리(연구비)',
                '연도별예산비교',
                '예산분석',
                '연도별예산데이터',
                '예산분석데이터'
            ]

            # 현재 시트 순서 확인
            current_sheet_names = workbook.sheetnames
            logging.info(f"현재 시트 순서: {current_sheet_names}")

            # 원하는 순서대로 시트 배치
            for target_index, sheet_name in enumerate(desired_order):
                if sheet_name in current_sheet_names:
                    # 현재 시트의 위치 찾기
                    current_index = workbook.sheetnames.index(sheet_name)
                    
                    # 시트가 원하는 위치에 있지 않다면 이동
                    if current_index != target_index:
                        sheet = workbook[sheet_name]
                        # 목표 위치로 이동 (offset 계산)
                        offset = target_index - current_index
                        workbook.move_sheet(sheet, offset=offset)
                        logging.debug(f"시트 '{sheet_name}' 이동: {current_index} -> {target_index}")
                else:
                    logging.debug(f"시트 '{sheet_name}'를 찾을 수 없습니다.")

            # 최종 시트 순서 로깅
            final_sheet_names = workbook.sheetnames
            logging.info(f"최종 시트 순서: {final_sheet_names}")

        except Exception as e:
            logging.error(f"시트 순서 조정 중 오류: {str(e)}")

    def _prepare_business_sheet(self, data: pd.DataFrame) -> pd.DataFrame:
        '''사업비 시트 데이터를 준비합니다.'''
        # 필요한 컬럼만 선택
        available_columns = [col for col in OUTPUT_COLUMNS if col in data.columns]

        if not available_columns:
            logging.warning("사업비 데이터에서 출력 가능한 컬럼을 찾을 수 없습니다.")
            return pd.DataFrame()

        result = data[available_columns].copy()

        # 누락된 컬럼은 빈 값으로 추가
        for col in OUTPUT_COLUMNS:
            if col not in result.columns:
                result[col] = ''
                logging.warning(f"사업비 데이터에 '{col}' 컬럼이 없어 빈 값으로 추가했습니다.")

        # 발의일자 컬럼 날짜 형식 처리
        result = self._format_date_columns(result)

        # 컬럼 순서 정렬
        result = result[OUTPUT_COLUMNS]

        return result

    def _prepare_research_sheet(self, data: pd.DataFrame) -> pd.DataFrame:
        '''연구비 시트 데이터를 준비합니다.'''
        # 기본 컬럼 처리
        available_columns = [col for col in OUTPUT_COLUMNS if col in data.columns]

        if not available_columns:
            logging.warning("연구비 데이터에서 출력 가능한 컬럼을 찾을 수 없습니다.")
            return pd.DataFrame()

        result = data[available_columns].copy()

        # 누락된 기본 컬럼은 빈 값으로 추가
        for col in OUTPUT_COLUMNS:
            if col not in result.columns:
                result[col] = ''
                logging.warning(f"연구비 데이터에 '{col}' 컬럼이 없어 빈 값으로 추가했습니다.")

        # 발의일자 컬럼 날짜 형식 처리
        result = self._format_date_columns(result)

        # 연구비 전용 추가 컬럼 추가
        result = self._add_research_specific_columns(result)

        # 최종 컬럼 순서 정렬
        final_columns = OUTPUT_COLUMNS + RESEARCH_ADDITIONAL_COLUMNS
        result = result[final_columns]

        return result

    def _add_research_specific_columns(self, data: pd.DataFrame) -> pd.DataFrame:
        '''연구비 전용 컬럼을 추가합니다.'''
        result = data.copy()

        # 연구자 정보 추출 (적요에서 _이름 형태로 추출)
        if '적요' in result.columns:
            result['연구자'] = result['적요'].apply(self._extract_researcher_name)
        else:
            result['연구자'] = ''
            logging.warning("적요 컬럼이 없어 연구자 정보를 추출할 수 없습니다.")

        # 반영일 컬럼은 빈 값으로 추가
        result['반영일'] = ''

        return result

    def _extract_researcher_name(self, summary_text: str) -> str:
        '''적요에서 연구자 이름을 추출합니다.'''
        if pd.isna(summary_text) or not isinstance(summary_text, str):
            return ''

        try:
            # _ 뒤에 나오는 한글을 추출
            import re
            pattern = r'_([가-힣]+)'
            match = re.search(pattern, summary_text)

            if match:
                return match.group(1)  # 한글 이름 부분만 반환
            else:
                return ''
        except Exception as e:
            logging.warning(f"연구자 이름 추출 중 오류: {str(e)}")
            return ''

    def _extract_research_topic(self, summary_text: str) -> str:
        '''적요에서 연구주제를 추출합니다.'''
        if pd.isna(summary_text) or not isinstance(summary_text, str):
            return ''

        try:
            # 25 심층연구(주제) 패턴에서 주제 부분 추출
            import re
            pattern = r'25 심층연구\(([^)]+)\)'
            match = re.search(pattern, summary_text)

            if match:
                topic = match.group(1)
                return topic  # 원본 주제명 그대로 반환
            else:
                return ''
        except Exception as e:
            logging.warning(f"연구주제 추출 중 오류: {str(e)} - 적요: {summary_text}")
            return ''

    def _format_date_columns(self, data: pd.DataFrame) -> pd.DataFrame:
        '''날짜 컬럼의 형식을 처리합니다 (시간 제거).'''
        result = data.copy()

        # 발의일자 컬럼이 있는 경우 날짜 형식 처리
        if '발의일자' in result.columns:
            try:
                # datetime 형식으로 변환 후 날짜만 추출
                result['발의일자'] = pd.to_datetime(result['발의일자'], errors='coerce')
                result['발의일자'] = result['발의일자'].dt.strftime('%Y-%m-%d')

                # NaT (Not a Time) 값은 빈 문자열로 처리
                result['발의일자'] = result['발의일자'].fillna('')

                logging.info("발의일자 컬럼 날짜 형식 처리 완료")
            except Exception as e:
                logging.warning(f"발의일자 컬럼 형식 처리 중 오류: {str(e)}")

        return result

    def _apply_sheet_styling(self, worksheet, data_frame, sheet_type):
        '''시트에 스타일링을 적용합니다.'''
        try:
            # 1. 컬럼 너비 조정
            self._apply_column_widths(worksheet, data_frame.columns)

            # 2. 헤더 스타일 적용
            self._apply_header_style(worksheet, data_frame.columns)

        except Exception as e:
            logging.error(f"시트 스타일링 적용 중 오류: {str(e)}")

    def _apply_column_widths(self, worksheet, columns):
        '''컬럼 너비를 조정합니다.'''
        from openpyxl.utils import get_column_letter

        for idx, column_name in enumerate(columns, 1):
            column_letter = get_column_letter(idx)
            width = EXCEL_STYLING['column_widths'].get(column_name, 12)  # 기본값 12
            worksheet.column_dimensions[column_letter].width = width

    def _apply_header_style(self, worksheet, columns):
        '''헤더 행에 스타일을 적용합니다.'''
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        # 헤더 스타일 정의
        header_fill = PatternFill(
            start_color=EXCEL_STYLING['header_style']['fill_color'],
            end_color=EXCEL_STYLING['header_style']['fill_color'],
            fill_type="solid"
        )
        header_font = Font(
            color=EXCEL_STYLING['header_style']['font_color'],
            bold=EXCEL_STYLING['header_style']['bold']
        )

        # 첫 번째 행(헤더)에 스타일 적용
        for idx in range(1, len(columns) + 1):
            column_letter = get_column_letter(idx)
            cell = worksheet[f'{column_letter}1']
            cell.fill = header_fill
            cell.font = header_font

    def _apply_summary_sheet_styling(self, worksheet, summary_data: pd.DataFrame):
        '''사업비 요약 시트에 특별한 스타일링을 적용합니다.'''
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # 컬럼 너비 설정
            column_widths = {
                '예산목': 15,
                '세목': 20,
                '예산과목': 25,
                '예산금액': 15,
                '지출액': 15,
                '예산잔액': 15,
                '집행률': 10
            }

            for idx, column_name in enumerate(SUMMARY_SHEET_COLUMNS, 1):
                column_letter = get_column_letter(idx)
                width = column_widths.get(column_name, 12)
                worksheet.column_dimensions[column_letter].width = width

            # 헤더 스타일
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type="solid")
            header_font = Font(color='FFFFFF', bold=True)

            # 테두리 스타일
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 헤더 행 스타일링
            for idx in range(1, len(SUMMARY_SHEET_COLUMNS) + 1):
                column_letter = get_column_letter(idx)
                cell = worksheet[f'{column_letter}1']
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # 예산목별 merge 범위 계산
            if self._is_research_summary_sheet(summary_data):
                # 연구비 시트: 복잡한 표 구조 처리
                budget_category_ranges = self._calculate_merge_ranges(summary_data)
                subcategory_ranges = self._calculate_subcategory_merge_ranges(summary_data)
            else:
                # 사업비 시트: 단순한 구조 처리
                budget_category_ranges = self._calculate_simple_merge_ranges(summary_data, '예산목')
                subcategory_ranges = self._calculate_simple_merge_ranges(summary_data, '세목')

            # 연구비 시트인 경우 총합 표에 Excel 함수 적용
            if self._is_research_summary_sheet(summary_data):
                budget_item_mapping = self._create_budget_item_mapping(summary_data)
                self._apply_total_summary_formulas(worksheet, summary_data, budget_item_mapping)

            # 예산목 컬럼 merge & center 적용 (A열)
            for budget_category, (start_row, end_row) in budget_category_ranges.items():
                if end_row > start_row:  # 여러 행에 걸쳐 있는 경우만 merge
                    merge_range = f'A{start_row + 2}:A{end_row + 2}'  # +2는 헤더 행 때문
                    worksheet.merge_cells(merge_range)
                    # merge된 셀에 중앙 정렬 적용
                    merged_cell = worksheet[f'A{start_row + 2}']
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    merged_cell.font = Font(bold=False)

            # 세목 컬럼 merge & center 적용 (B열)
            for subcategory_key, (start_row, end_row) in subcategory_ranges.items():
                if end_row > start_row:  # 여러 행에 걸쳐 있는 경우만 merge
                    merge_range = f'B{start_row + 2}:B{end_row + 2}'  # +2는 헤더 행 때문
                    worksheet.merge_cells(merge_range)
                    # merge된 셀에 중앙 정렬 적용
                    merged_cell = worksheet[f'B{start_row + 2}']
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    merged_cell.font = Font(bold=False)

            # Excel 함수 적용 및 데이터 행 스타일링
            total_row_idx = None
            for row_idx in range(2, len(summary_data) + 2):
                row_data = summary_data.iloc[row_idx - 2]

                # 총액 행 인덱스 찾기
                if row_data['예산목'] == '총액':
                    total_row_idx = row_idx

                for col_idx in range(1, len(SUMMARY_SHEET_COLUMNS) + 1):
                    column_letter = get_column_letter(col_idx)
                    cell = worksheet[f'{column_letter}{row_idx}']
                    cell.border = thin_border

                    # Excel 함수 적용
                    if row_data['예산목'] == '총액':
                        # 총액 행의 함수 적용
                        if col_idx == 4:  # 예산금액 총합
                            cell.value = f'=SUM(D2:D{row_idx-1})'
                        elif col_idx == 5:  # 지출액 총합
                            cell.value = f'=SUM(E2:E{row_idx-1})'
                        elif col_idx == 6:  # 예산잔액 총합
                            cell.value = f'=SUM(F2:F{row_idx-1})'
                        elif col_idx == 7:  # 총 집행률
                            cell.value = f'=IF(D{row_idx}=0,0,ROUND(E{row_idx}/D{row_idx}*100,0))'
                    else:
                        # 일반 행의 함수 적용
                        if col_idx == 6:  # 예산잔액 = 예산금액 - 지출액
                            cell.value = f'=D{row_idx}-E{row_idx}'
                        elif col_idx == 7:  # 집행률 = 지출액/예산금액*100
                            cell.value = f'=IF(D{row_idx}=0,0,ROUND(E{row_idx}/D{row_idx}*100,0))'

                    # 총액 행 강조
                    if row_data['예산목'] == '총액':
                        cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type="solid")
                        cell.font = Font(bold=True, color='FF0000')

                    # 표 제목 행 강조 (총합, 연구주제명 등)
                    elif (row_data['예산목'] == '총합' or
                          (row_data['예산목'] and row_data['세목'] and not row_data['예산과목'])):
                        cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type="solid")
                        cell.font = Font(bold=True, color='1F4E79')
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # 숫자 컬럼 우측 정렬 및 숫자 포맷팅
                    if col_idx in [4, 5, 6]:  # 예산금액, 지출액, 예산잔액
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        # 천단위 구분자 적용
                        cell.number_format = '#,##0'
                    elif col_idx == 7:  # 집행률
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

            logging.info("사업비 요약 시트 스타일링 적용 완료")

        except Exception as e:
            logging.error(f"사업비 요약 시트 스타일링 적용 중 오류: {str(e)}")

    def _apply_total_sheet_styling(self, worksheet, total_data: pd.DataFrame):
        '''총액 시트에 스타일링을 적용합니다.'''
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # 컬럼 너비 설정 (총액 시트용)
            column_widths = {
                '예산목': 15,
                '세목': 20,
                '예산과목': 25,
                '예산금액': 15,
                '센터': 15,
                '심층연구': 15,
                '예산잔액': 15,
                '집행률': 12
            }

            for idx, column in enumerate(TOTAL_SHEET_COLUMNS, 1):
                column_letter = get_column_letter(idx)
                worksheet.column_dimensions[column_letter].width = column_widths.get(column, 12)

            # 스타일 정의
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 헤더 행 스타일링
            for idx in range(1, len(TOTAL_SHEET_COLUMNS) + 1):
                column_letter = get_column_letter(idx)
                cell = worksheet[f'{column_letter}1']
                cell.fill = header_fill
                cell.font = header_font

            # 예산목별 merge 범위 계산 (총액 시트용)
            budget_category_ranges = self._calculate_total_merge_ranges(total_data, '예산목')
            subcategory_ranges = self._calculate_total_merge_ranges(total_data, '세목')

            # 예산목 컬럼 merge & center 적용 (A열)
            for _, (start_row, end_row) in budget_category_ranges.items():
                if end_row > start_row:
                    merge_range = f'A{start_row + 2}:A{end_row + 2}'
                    worksheet.merge_cells(merge_range)
                    merged_cell = worksheet[f'A{start_row + 2}']
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    merged_cell.font = Font(bold=False)

            # 세목 컬럼 merge & center 적용 (B열)
            for _, (start_row, end_row) in subcategory_ranges.items():
                if end_row > start_row:
                    merge_range = f'B{start_row + 2}:B{end_row + 2}'
                    worksheet.merge_cells(merge_range)
                    merged_cell = worksheet[f'B{start_row + 2}']
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    merged_cell.font = Font(bold=False)

            # Excel 함수 적용 및 데이터 행 스타일링
            for row_idx in range(2, len(total_data) + 2):
                row_data = total_data.iloc[row_idx - 2]

                for col_idx in range(1, len(TOTAL_SHEET_COLUMNS) + 1):
                    column_letter = get_column_letter(col_idx)
                    cell = worksheet[f'{column_letter}{row_idx}']
                    cell.border = thin_border

                    # Excel 함수 적용
                    if row_data['예산목'] == '총액':
                        # 총액 행의 함수 적용
                        if col_idx == 4:  # 예산금액 총합
                            cell.value = f'=SUM(D2:D{row_idx-1})'
                        elif col_idx == 5:  # 센터 총합
                            cell.value = f'=SUM(E2:E{row_idx-1})'
                        elif col_idx == 6:  # 심층연구 총합
                            cell.value = f'=SUM(F2:F{row_idx-1})'
                        elif col_idx == 7:  # 예산잔액 총합
                            cell.value = f'=SUM(G2:G{row_idx-1})'
                        elif col_idx == 8:  # 총 집행률
                            cell.value = f'=IF(D{row_idx}=0,0,ROUND((E{row_idx}+F{row_idx})/D{row_idx}*100,0))'
                    else:
                        # 일반 행의 함수 적용
                        if col_idx == 7:  # 예산잔액 = 예산금액 - 센터 - 심층연구
                            cell.value = f'=D{row_idx}-E{row_idx}-F{row_idx}'
                        elif col_idx == 8:  # 집행률 = (센터+심층연구)/예산금액*100
                            cell.value = f'=IF(D{row_idx}=0,0,ROUND((E{row_idx}+F{row_idx})/D{row_idx}*100,0))'

                    # 셀 정렬 및 형식 설정
                    if col_idx in [4, 5, 6, 7]:  # 금액 컬럼들
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '#,##0'
                    elif col_idx == 8:  # 집행률
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

            logging.info("총액 시트 스타일링 적용 완료")

        except Exception as e:
            logging.error(f"총액 시트 스타일링 적용 중 오류: {str(e)}")

    def _calculate_total_merge_ranges(self, total_data: pd.DataFrame, column: str) -> dict:
        '''총액 시트의 merge 범위를 계산합니다. (총액 행 제외)'''
        merge_ranges = {}
        current_value = None
        start_row = None

        for idx, row in total_data.iterrows():
            value = row[column]

            # 총액 행은 merge 대상에서 제외
            if value == '총액':
                if current_value and start_row is not None:
                    merge_ranges[f"{current_value}_{start_row}"] = (start_row, idx - 1)
                current_value = None
                start_row = None
                continue

            if value and value != '':
                if current_value != value:
                    if current_value and start_row is not None:
                        merge_ranges[f"{current_value}_{start_row}"] = (start_row, idx - 1)
                    current_value = value
                    start_row = idx

        # 마지막 값 처리 (총액 행 제외)
        if current_value and start_row is not None:
            # 총액 행이 있다면 그 전까지만 merge
            last_row = len(total_data) - 1
            if total_data.iloc[last_row]['예산목'] == '총액':
                last_row -= 1
            merge_ranges[f"{current_value}_{start_row}"] = (start_row, last_row)

        return merge_ranges

    def _calculate_merge_ranges(self, summary_data: pd.DataFrame) -> dict:
        '''예산목별 merge 범위를 계산합니다. (표별 독립 처리)'''
        merge_ranges = {}

        # 표 경계 식별
        table_boundaries = self._identify_table_boundaries(summary_data)
        logging.debug(f"Merge 계산용 표 경계: {table_boundaries}")

        # 각 표별로 독립적으로 merge 범위 계산
        for table_name, table_start, table_end in table_boundaries:
            logging.debug(f"표 '{table_name}' merge 범위 계산 중 ({table_start}~{table_end})")

            current_budget_category = None
            start_row = None

            for idx in range(table_start, table_end + 1):
                if idx >= len(summary_data):
                    break

                row = summary_data.iloc[idx]
                budget_category = row['예산목']

                logging.debug(f"  행 {idx}: 예산목='{budget_category}'")

                # 빈 행이나 총액 행 건너뛰기
                if (self._is_completely_empty_row(row) or
                    budget_category == '총액'):
                    if budget_category == '총액' and current_budget_category and start_row is not None:
                        # 총액 행 전까지 merge
                        range_key = f"{current_budget_category}_{table_name}_{start_row}"
                        merge_ranges[range_key] = (start_row, idx - 1)
                        logging.debug(f"    총액 전까지 merge: {range_key} = ({start_row}, {idx-1})")
                        current_budget_category = None
                        start_row = None
                    continue

                if budget_category and budget_category != '':  # 새로운 예산목 시작
                    if current_budget_category and start_row is not None:
                        # 이전 예산목의 범위 저장
                        range_key = f"{current_budget_category}_{table_name}_{start_row}"
                        merge_ranges[range_key] = (start_row, idx - 1)
                        logging.debug(f"    이전 예산목 merge: {range_key} = ({start_row}, {idx-1})")

                    current_budget_category = budget_category
                    start_row = idx
                    logging.debug(f"    새 예산목 시작: {budget_category}, start_row={start_row}")

            # 표의 마지막 예산목 처리
            if current_budget_category and start_row is not None:
                range_key = f"{current_budget_category}_{table_name}_{start_row}"
                merge_ranges[range_key] = (start_row, table_end)
                logging.debug(f"    마지막 예산목 merge: {range_key} = ({start_row}, {table_end})")

        logging.info(f"계산된 예산목 merge 범위: {len(merge_ranges)}개")
        return merge_ranges

    def _calculate_simple_merge_ranges(self, summary_data: pd.DataFrame, column: str) -> dict:
        '''사업비 시트용 단순한 merge 범위를 계산합니다.'''
        merge_ranges = {}
        current_value = None
        start_row = None

        for idx, row in summary_data.iterrows():
            value = row[column]

            # 총액 행은 제외
            if value == '총액':
                if current_value and start_row is not None:
                    merge_ranges[f"{current_value}_{start_row}"] = (start_row, idx - 1)
                current_value = None
                start_row = None
                continue

            if value and value != '':
                if current_value != value:
                    if current_value and start_row is not None:
                        merge_ranges[f"{current_value}_{start_row}"] = (start_row, idx - 1)
                    current_value = value
                    start_row = idx

        # 마지막 값 처리
        if current_value and start_row is not None:
            merge_ranges[f"{current_value}_{start_row}"] = (start_row, len(summary_data) - 2)  # 총액 행 제외

        return merge_ranges

    def _identify_table_boundaries(self, summary_data: pd.DataFrame) -> list:
        '''
        DataFrame에서 각 표의 경계를 식별합니다.

        Returns:
            list: [(table_name, start_idx, end_idx), ...] 형태의 표 경계 정보
        '''
        try:
            boundaries = []
            current_table = None
            start_idx = None

            logging.debug("표 경계 식별 시작")
            logging.debug(f"전체 데이터 크기: {len(summary_data)}")

            for idx, row in summary_data.iterrows():
                budget_category = row['예산목']
                subcategory = row['세목']
                budget_item = row['예산과목']

                logging.debug(f"행 {idx}: '{budget_category}' | '{subcategory}' | '{budget_item}'")

                # 표 제목 행 감지 (예: "총합", "AI 박상헌" 등)
                is_title_row = (budget_category in ['총합'] or
                               (budget_category and subcategory and
                                (pd.isna(budget_item) or budget_item == '')))

                if is_title_row:
                    logging.debug(f"표 제목 행 발견: {budget_category} {subcategory}")

                    # 이전 표 경계 저장
                    if current_table and start_idx is not None:
                        # 이전 표의 끝을 찾음 (현재 제목 행 직전까지)
                        end_idx = self._find_table_end(summary_data, start_idx, idx - 1)
                        boundaries.append((current_table, start_idx, end_idx))
                        logging.debug(f"이전 표 저장: {current_table} ({start_idx}~{end_idx})")

                    # 새 표 시작
                    if budget_category == '총합':
                        current_table = '총합'
                    else:
                        current_table = f"{budget_category} {subcategory}"
                    start_idx = idx + 1  # 제목 다음 행부터 시작
                    logging.debug(f"새 표 시작: {current_table}, start_idx={start_idx}")

                elif budget_category == '총액':
                    # 총합 표의 끝
                    if current_table == '총합' and start_idx is not None:
                        boundaries.append((current_table, start_idx, idx - 1))
                        logging.debug(f"총합 표 완료: {current_table} ({start_idx}~{idx-1})")
                        current_table = None
                        start_idx = None

            # 마지막 표 처리
            if current_table and start_idx is not None:
                end_idx = self._find_table_end(summary_data, start_idx, len(summary_data) - 1)
                boundaries.append((current_table, start_idx, end_idx))
                logging.debug(f"마지막 표 저장: {current_table} ({start_idx}~{end_idx})")

            logging.info(f"식별된 표 경계: {[(name, start, end) for name, start, end in boundaries]}")
            return boundaries

        except Exception as e:
            logging.error(f"표 경계 식별 중 오류: {str(e)}")
            return []

    def _find_table_end(self, summary_data: pd.DataFrame, start_idx: int, max_idx: int) -> int:
        '''
        표의 실제 끝 인덱스를 찾습니다. (빈 행과 다음 표 제목 행 제외)
        '''
        try:
            end_idx = start_idx
            logging.debug(f"표 끝 찾기 시작: start_idx={start_idx}, max_idx={max_idx}")

            for idx in range(start_idx, min(max_idx + 1, len(summary_data))):
                row = summary_data.iloc[idx]
                logging.debug(f"행 {idx}: 예산목='{row['예산목']}', 세목='{row['세목']}', 예산과목='{row['예산과목']}'")

                # 완전히 빈 행이면 표의 끝으로 간주
                if self._is_completely_empty_row(row):
                    logging.debug(f"빈 행 발견으로 표 끝: {idx-1}")
                    break

                # 다음 표의 제목 행이면 표의 끝 (예: "AI 박상헌")
                if (row['예산목'] and row['세목'] and
                    (pd.isna(row['예산과목']) or row['예산과목'] == '')):
                    logging.debug(f"다음 표 제목 행 발견으로 표 끝: {idx-1}")
                    break

                # 실제 데이터가 있는 행이면 계속
                if row['예산과목'] and row['예산과목'] != '':
                    end_idx = idx
                    logging.debug(f"데이터 행 발견, end_idx 업데이트: {end_idx}")

            logging.debug(f"최종 표 끝 인덱스: {end_idx}")
            return end_idx

        except Exception as e:
            logging.error(f"표 끝 찾기 중 오류: {str(e)}")
            return start_idx

    def _is_completely_empty_row(self, row: pd.Series) -> bool:
        '''
        행이 완전히 비어있는지 확인합니다.
        '''
        return ((pd.isna(row['예산목']) or row['예산목'] == '') and
                (pd.isna(row['세목']) or row['세목'] == '') and
                (pd.isna(row['예산과목']) or row['예산과목'] == ''))

    def _calculate_subcategory_merge_ranges(self, summary_data: pd.DataFrame) -> dict:
        '''세목별 merge 범위를 계산합니다. (표별 독립 처리)'''
        merge_ranges = {}

        # 표 경계 식별
        table_boundaries = self._identify_table_boundaries(summary_data)
        logging.debug(f"세목 Merge 계산용 표 경계: {table_boundaries}")

        # 각 표별로 독립적으로 merge 범위 계산
        for table_name, table_start, table_end in table_boundaries:
            logging.debug(f"표 '{table_name}' 세목 merge 범위 계산 중 ({table_start}~{table_end})")

            current_subcategory = None
            start_row = None

            for idx in range(table_start, table_end + 1):
                if idx >= len(summary_data):
                    break

                row = summary_data.iloc[idx]
                subcategory = row['세목']
                budget_category = row['예산목']

                logging.debug(f"  행 {idx}: 세목='{subcategory}', 예산목='{budget_category}'")

                # 빈 행이나 총액 행 건너뛰기
                if (self._is_completely_empty_row(row) or
                    budget_category == '총액'):
                    if budget_category == '총액' and current_subcategory and start_row is not None:
                        # 총액 행 전까지 merge
                        range_key = f"{current_subcategory}_{table_name}_{start_row}"
                        merge_ranges[range_key] = (start_row, idx - 1)
                        logging.debug(f"    총액 전까지 세목 merge: {range_key} = ({start_row}, {idx-1})")
                        current_subcategory = None
                        start_row = None
                    continue

                if subcategory and subcategory != '':  # 새로운 세목 시작
                    if current_subcategory and start_row is not None:
                        # 이전 세목의 범위 저장
                        range_key = f"{current_subcategory}_{table_name}_{start_row}"
                        merge_ranges[range_key] = (start_row, idx - 1)
                        logging.debug(f"    이전 세목 merge: {range_key} = ({start_row}, {idx-1})")

                    current_subcategory = subcategory
                    start_row = idx
                    logging.debug(f"    새 세목 시작: {subcategory}, start_row={start_row}")

            # 표의 마지막 세목 처리
            if current_subcategory and start_row is not None:
                range_key = f"{current_subcategory}_{table_name}_{start_row}"
                merge_ranges[range_key] = (start_row, table_end)
                logging.debug(f"    마지막 세목 merge: {range_key} = ({start_row}, {table_end})")

        logging.info(f"계산된 세목 merge 범위: {len(merge_ranges)}개")
        return merge_ranges

    def _create_budget_item_mapping(self, summary_data: pd.DataFrame) -> dict:
        '''
        각 표에서 예산과목별 행 번호를 매핑합니다.

        Returns:
            dict: {table_name: {budget_item: row_number}} 형태의 매핑
        '''
        try:
            mapping = {}
            table_boundaries = self._identify_table_boundaries(summary_data)

            for table_name, table_start, table_end in table_boundaries:
                table_mapping = {}

                for idx in range(table_start, table_end + 1):
                    if idx >= len(summary_data):
                        break

                    row = summary_data.iloc[idx]
                    budget_item = row['예산과목']

                    # 실제 예산과목이 있는 행만 매핑
                    if budget_item and budget_item != '' and not pd.isna(budget_item):
                        # Excel 행 번호는 1-based이고 헤더가 있으므로 +2
                        excel_row = idx + 2
                        table_mapping[budget_item] = excel_row

                mapping[table_name] = table_mapping

            logging.info(f"예산과목별 행 매핑 완료: {len(mapping)}개 표")
            return mapping

        except Exception as e:
            logging.error(f"예산과목별 행 매핑 중 오류: {str(e)}")
            return {}

    def _apply_total_summary_formulas(self, worksheet, summary_data: pd.DataFrame,
                                    budget_item_mapping: dict):
        '''
        총합 표에 개별 표들을 참조하는 Excel 함수를 적용합니다.
        '''
        try:
            from openpyxl.utils import get_column_letter

            # 총합 표 영역 식별 (표 제목 행 제외)
            total_table_start = None
            total_table_end = None

            for idx, row in summary_data.iterrows():
                if row['예산목'] == '총합':
                    total_table_start = idx + 1  # 제목 다음 행부터
                elif row['예산목'] == '총액' and total_table_start is not None:
                    total_table_end = idx - 1  # 총액 행 전까지
                    break

            if total_table_start is None or total_table_end is None:
                logging.warning("총합 표 영역을 찾을 수 없습니다.")
                return

            # 개별 표들의 이름 추출 (총합 제외)
            individual_tables = [name for name in budget_item_mapping.keys()
                               if name != '총합']

            # 총합 표의 각 예산과목에 대해 Excel 함수 적용
            for idx in range(total_table_start, total_table_end + 1):
                if idx >= len(summary_data):
                    break

                row = summary_data.iloc[idx]
                budget_item = row['예산과목']

                if budget_item and budget_item != '' and not pd.isna(budget_item):
                    excel_row = idx + 2  # Excel 행 번호

                    # 예산금액(D열)과 지출액(E열)에 SUM 함수 적용
                    budget_formula = self._create_sum_formula(
                        budget_item, individual_tables, budget_item_mapping, 'D')
                    expense_formula = self._create_sum_formula(
                        budget_item, individual_tables, budget_item_mapping, 'E')

                    if budget_formula:
                        worksheet[f'D{excel_row}'].value = budget_formula
                    if expense_formula:
                        worksheet[f'E{excel_row}'].value = expense_formula

            logging.info("총합 표 Excel 함수 적용 완료")

        except Exception as e:
            logging.error(f"총합 표 Excel 함수 적용 중 오류: {str(e)}")

    def _create_sum_formula(self, budget_item: str, individual_tables: list,
                          budget_item_mapping: dict, column: str) -> str:
        '''
        특정 예산과목에 대한 SUM 함수를 생성합니다.
        '''
        try:
            cell_references = []

            for table_name in individual_tables:
                table_mapping = budget_item_mapping.get(table_name, {})
                if budget_item in table_mapping:
                    row_number = table_mapping[budget_item]
                    cell_references.append(f'{column}{row_number}')

            if cell_references:
                return f'=SUM({",".join(cell_references)})'
            else:
                return None

        except Exception as e:
            logging.error(f"SUM 함수 생성 중 오류: {str(e)}")
            return None

    def _is_research_summary_sheet(self, summary_data: pd.DataFrame) -> bool:
        '''
        연구비 요약 시트인지 확인합니다. (총합 표와 개별 표들이 있는지 확인)
        '''
        try:
            # '총합' 제목 행이 있는지 확인
            has_total_summary = any(row['예산목'] == '총합' for _, row in summary_data.iterrows())

            # 개별 표 제목 행이 있는지 확인 (예산목과 세목은 있지만 예산과목은 없는 행)
            has_individual_tables = any(
                (row['예산목'] and row['세목'] and
                 (pd.isna(row['예산과목']) or row['예산과목'] == ''))
                for _, row in summary_data.iterrows()
                if row['예산목'] != '총합'
            )

            return has_total_summary and has_individual_tables

        except Exception as e:
            logging.error(f"연구비 시트 확인 중 오류: {str(e)}")
            return False


class SummarySheetGenerator:
    '''사업비 요약 시트 생성 클래스'''

    def __init__(self):
        self.budget_classification = BUDGET_CLASSIFICATION
        self.summary_data = None

    def _extract_research_topic(self, summary_text: str) -> str:
        '''적요에서 연구주제를 추출합니다.'''
        if pd.isna(summary_text) or not isinstance(summary_text, str):
            return ''

        try:
            # 25 심층연구(주제) 패턴에서 주제 부분 추출
            import re
            pattern = r'25 심층연구\(([^)]+)\)'
            match = re.search(pattern, summary_text)

            if match:
                topic = match.group(1)
                return topic  # 원본 주제명 그대로 반환
            else:
                return ''
        except Exception as e:
            logging.warning(f"연구주제 추출 중 오류: {str(e)} - 적요: {summary_text}")
            return ''

    def _extract_researcher_name(self, summary_text: str) -> str:
        '''적요에서 연구자 이름을 추출합니다.'''
        if pd.isna(summary_text) or not isinstance(summary_text, str):
            return ''

        try:
            # _ 뒤에 나오는 한글을 추출
            import re
            pattern = r'_([가-힣]+)'
            match = re.search(pattern, summary_text)

            if match:
                return match.group(1)  # 한글 이름 부분만 반환
            else:
                return ''
        except Exception as e:
            logging.warning(f"연구자 이름 추출 중 오류: {str(e)}")
            return ''

    def generate_summary_sheet(self, business_data: pd.DataFrame) -> pd.DataFrame:
        '''
        집행관리 데이터에서 사업비 요약 시트를 생성합니다.

        Args:
            business_data (pd.DataFrame): 사업비 집행관리 데이터

        Returns:
            pd.DataFrame: 사업비 요약 시트 데이터
        '''
        if business_data is None or business_data.empty:
            logging.warning("사업비 데이터가 비어있습니다.")
            return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

        try:
            logging.info("사업비 요약 시트 생성 시작")

            # 1. 예산과목별 지출액 집계
            expense_summary = self._aggregate_expenses(business_data)

            # 2. 계층적 구조로 변환
            hierarchical_data = self._create_hierarchical_structure(expense_summary)

            # 3. 예산금액 및 계산 컬럼 추가
            final_data = self._add_budget_calculations(hierarchical_data)

            self.summary_data = final_data
            logging.info(f"사업비 요약 시트 생성 완료: {len(final_data)}건")

            return final_data

        except Exception as e:
            logging.error(f"사업비 요약 시트 생성 중 오류: {str(e)}")
            return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

    def _aggregate_expenses(self, business_data: pd.DataFrame) -> pd.DataFrame:
        '''예산과목별 지출액을 집계합니다.'''
        try:
            # 필요한 컬럼 확인
            required_columns = ['예산과목', '총지급액']
            missing_columns = [col for col in required_columns if col not in business_data.columns]

            if missing_columns:
                logging.error(f"필수 컬럼이 없습니다: {missing_columns}")
                return pd.DataFrame()

            # 데이터 전처리
            work_data = business_data.copy()
            work_data['예산과목'] = work_data['예산과목'].fillna('미분류').astype(str)
            work_data['총지급액'] = pd.to_numeric(work_data['총지급액'], errors='coerce').fillna(0)

            # 예산과목별 지출액 집계
            expense_summary = work_data.groupby('예산과목')['총지급액'].sum().reset_index()
            expense_summary.columns = ['예산과목', '지출액']

            logging.info(f"예산과목별 집계 완료: {len(expense_summary)}개 항목")
            return expense_summary

        except Exception as e:
            logging.error(f"지출액 집계 중 오류: {str(e)}")
            return pd.DataFrame()

    def _create_hierarchical_structure(self, expense_summary: pd.DataFrame) -> pd.DataFrame:
        '''집계된 데이터를 flat 구조로 변환.'''
        try:
            flat_rows = []
            budget_categories = self.budget_classification['budget_categories']
            total_expense = 0

            # 각 예산목별로 처리
            for budget_category, category_info in budget_categories.items():
                is_first_category_item = True

                # 각 세목별로 처리
                for subcategory, budget_items in category_info['subcategories'].items():
                    is_first_subcategory_item = True

                    # 각 예산과목별로 처리
                    for budget_item in budget_items:
                        # 실제 지출액 찾기 (정확한 매칭 우선, 그 다음 부분 매칭)
                        exact_match = expense_summary[expense_summary['예산과목'] == budget_item]
                        if not exact_match.empty:
                            matching_rows = exact_match
                        else:
                            # 정규식 특수문자 이스케이프 처리
                            escaped_item = budget_item.replace('(', r'\(').replace(')', r'\)')
                            matching_rows = expense_summary[
                                expense_summary['예산과목'].str.contains(escaped_item, na=False, case=False, regex=True)
                            ]

                        if not matching_rows.empty:
                            item_expense = matching_rows['지출액'].sum()
                            # 중복 방지를 위해 사용된 행들을 제거
                            expense_summary = expense_summary.drop(matching_rows.index)
                        else:
                            item_expense = 0

                        # 행 데이터 생성
                        row_data = {
                            '예산목': budget_category if is_first_category_item else '',
                            '세목': subcategory if is_first_subcategory_item else '',
                            '예산과목': budget_item,
                            '지출액': item_expense
                        }

                        flat_rows.append(row_data)
                        total_expense += item_expense

                        # 플래그 업데이트
                        is_first_category_item = False
                        is_first_subcategory_item = False

            # 총액 행 추가
            flat_rows.append({
                '예산목': '총액',
                '세목': '',
                '예산과목': '',
                '지출액': total_expense
            })

            result_df = pd.DataFrame(flat_rows)

            return result_df

        except Exception as e:
            logging.error(f"flat 구조 생성 중 오류: {str(e)}")
            return pd.DataFrame()

    def _add_budget_calculations(self, hierarchical_data: pd.DataFrame) -> pd.DataFrame:
        '''예산금액, 예산잔액, 집행률을 계산하여 추가합니다.'''
        try:
            if hierarchical_data.empty:
                return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

            result_data = hierarchical_data.copy()
            default_budgets = self.budget_classification['2025_budget_amounts']

            # 예산금액 설정
            result_data['예산금액'] = result_data['예산과목'].apply(
                lambda x: self._get_budget_amount(x, default_budgets)
            )

            # 예산잔액 계산 (예산금액 - 지출액)
            result_data['예산잔액'] = result_data['예산금액'] - result_data['지출액']

            # 집행률 계산 (지출액 / 예산금액 * 100)
            result_data['집행률'] = result_data.apply(
                lambda row: f"{(row['지출액'] / row['예산금액'] * 100):.0f}"
                if row['예산금액'] > 0 else "0", axis=1
            )

            # 컬럼 순서 정렬
            result_data = result_data[SUMMARY_SHEET_COLUMNS]

            # 숫자 컬럼의 데이터 타입을 명시적으로 숫자로 변환
            number_columns = ['예산금액', '지출액', '예산잔액']
            for col in number_columns:
                if col in result_data.columns:
                    result_data[col] = pd.to_numeric(result_data[col], errors='coerce').fillna(0)

            logging.info("예산 계산 완료")
            return result_data

        except Exception as e:
            logging.error(f"예산 계산 중 오류: {str(e)}")
            return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

    def _get_budget_amount(self, budget_item: str, default_budgets: dict) -> int:
        '''예산과목에 해당하는 예산금액을 반환합니다.'''
        # 총액인 경우
        if budget_item == '총액' or budget_item == '':
            return sum(default_budgets.values())

        # 정확한 매칭 시도
        if budget_item in default_budgets:
            return default_budgets[budget_item]

        # 부분 매칭 시도
        for key, amount in default_budgets.items():
            if key in budget_item or budget_item in key:
                return amount

        return 0


    def generate_research_summary_sheet(self, research_data: pd.DataFrame) -> pd.DataFrame:
        '''
        집행관리 데이터에서 연구비 요약 시트를 생성합니다.
        총합 표와 연구주제/연구자별 개별 표들을 포함합니다.

        Args:
            research_data (pd.DataFrame): 연구비 집행관리 데이터

        Returns:
            pd.DataFrame: 연구비 요약 시트 데이터
        '''
        if research_data is None or research_data.empty:
            logging.warning("연구비 데이터가 비어있습니다.")
            return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

        try:
            logging.info("연구비 요약 시트 생성 시작")
            logging.info(f"연구비 데이터 크기: {research_data.shape}")
            logging.info(f"연구비 데이터 컬럼: {list(research_data.columns)}")

            # 적요 컬럼 샘플 확인
            if '적요' in research_data.columns:
                sample_summaries = research_data['적요'].head(10).tolist()
                logging.info(f"적요 샘플: {sample_summaries}")
            else:
                logging.error("적요 컬럼이 연구비 데이터에 없습니다!")

            # 1. 총합 표 생성
            total_summary = self._generate_total_summary(research_data)
            logging.info(f"총합 표 크기: {total_summary.shape}")

            # 2. 연구주제/연구자별 개별 표들 생성
            individual_summaries = self._generate_individual_summaries(research_data)
            logging.info(f"개별 표 개수: {len(individual_summaries)}")

            # 3. 총합 표와 개별 표들을 결합
            all_tables = [total_summary] + individual_summaries
            result_data = pd.concat(all_tables, ignore_index=True)

            logging.info(f"연구비 요약 시트 생성 완료: {len(result_data)}건")
            return result_data

        except Exception as e:
            logging.error(f"연구비 요약 시트 생성 중 오류: {str(e)}")
            return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

    def _generate_total_summary(self, research_data: pd.DataFrame) -> pd.DataFrame:
        '''총합 표를 생성합니다.'''
        try:
            # 1. 사업비와 동일한 구조로 기본 데이터 생성
            business_summary = self.generate_summary_sheet(research_data)

            if business_summary.empty:
                return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

            # 2. 총액 행을 찾아서 그 앞에 연구개발비와 유형자산 행 추가
            total_row_idx = None
            for idx, row in business_summary.iterrows():
                if row['예산목'] == '총액':
                    total_row_idx = idx
                    break

            if total_row_idx is not None:
                # 총액 행을 제거하고 새로운 행들을 추가
                result_data = business_summary.iloc[:total_row_idx].copy()

                # 연구개발비와 유형자산의 실제 지출액 계산
                research_dev_expense = self._calculate_research_expense(research_data, '연구개발비')
                asset_expense = self._calculate_research_expense(research_data, '자산취득비')

                # 연구개발비 행 추가
                research_dev_row = {
                    '예산목': '연구개발비',
                    '세목': '연구개발비',
                    '예산과목': '연구개발비',
                    '예산금액': 0,
                    '지출액': research_dev_expense,
                    '예산잔액': -research_dev_expense,
                    '집행률': '0%' if research_dev_expense == 0 else '∞%'
                }

                # 유형자산 행 추가
                asset_row = {
                    '예산목': '유형자산',
                    '세목': '자산취득비',
                    '예산과목': '자산취득비',
                    '예산금액': 0,
                    '지출액': asset_expense,
                    '예산잔액': -asset_expense,
                    '집행률': '0%' if asset_expense == 0 else '∞%'
                }

                # 새로운 행들을 DataFrame에 추가
                result_data = pd.concat([
                    result_data,
                    pd.DataFrame([research_dev_row]),
                    pd.DataFrame([asset_row])
                ], ignore_index=True)

                # 총액 행 다시 계산하여 추가
                total_expense = (result_data['지출액'].sum() if len(result_data) > 0 else 0)
                total_row = {
                    '예산목': '총액',
                    '세목': '',
                    '예산과목': '',
                    '예산금액': 0,
                    '지출액': total_expense,
                    '예산잔액': -total_expense,
                    '집행률': '0%' if total_expense == 0 else '∞%'
                }
                result_data = pd.concat([result_data, pd.DataFrame([total_row])], ignore_index=True)

            else:
                result_data = business_summary.copy()

            # 총합 표 제목 추가
            title_row = {
                '예산목': '총합',
                '세목': '',
                '예산과목': '',
                '예산금액': '',
                '지출액': '',
                '예산잔액': '',
                '집행률': ''
            }
            result_data = pd.concat([pd.DataFrame([title_row]), result_data], ignore_index=True)

            # 총합 표 뒤에 2개 빈 행 추가
            empty_rows = []
            for _ in range(2):
                empty_row = {col: '' for col in SUMMARY_SHEET_COLUMNS}
                empty_rows.append(empty_row)
            result_data = pd.concat([result_data, pd.DataFrame(empty_rows)], ignore_index=True)

            return result_data

        except Exception as e:
            logging.error(f"총합 표 생성 중 오류: {str(e)}")
            return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

    def _generate_individual_summaries(self, research_data: pd.DataFrame) -> list:
        '''연구주제/연구자별 개별 표들을 생성합니다.'''
        try:
            individual_summaries = []

            # 연구주제와 연구자 조합 추출
            topic_researcher_combinations = self._extract_topic_researcher_combinations(research_data)
            logging.info(f"추출된 연구주제/연구자 조합: {topic_researcher_combinations}")

            for topic, researcher in topic_researcher_combinations:
                logging.info(f"개별 표 생성 중: {topic} - {researcher}")

                # 해당 주제/연구자의 데이터만 필터링
                filtered_data = self._filter_data_by_topic_researcher(research_data, topic, researcher)
                logging.info(f"필터링된 데이터 건수: {len(filtered_data)}")

                if not filtered_data.empty:
                    # 개별 표 생성 (총액 행 없음)
                    individual_table = self._generate_individual_table(filtered_data, topic, researcher)
                    individual_summaries.append(individual_table)
                    logging.info(f"개별 표 생성 완료: {topic} - {researcher}")
                else:
                    logging.warning(f"필터링된 데이터가 비어있음: {topic} - {researcher}")

            logging.info(f"총 {len(individual_summaries)}개의 개별 표 생성됨")
            return individual_summaries

        except Exception as e:
            logging.error(f"개별 표 생성 중 오류: {str(e)}")
            return []

    def _extract_topic_researcher_combinations(self, research_data: pd.DataFrame) -> list:
        '''연구주제와 연구자 조합을 추출합니다.'''
        combinations = set()

        if '적요' in research_data.columns:
            logging.info(f"적요 컬럼에서 연구주제/연구자 조합 추출 시작 (총 {len(research_data)}건)")

            for idx, row in research_data.iterrows():
                summary = row['적요']
                topic = self._extract_research_topic(summary)
                researcher = self._extract_researcher_name(summary)

                if idx < 5:  # 처음 5개만 로그 출력
                    logging.info(f"적요: {summary}")
                    logging.info(f"추출된 주제: '{topic}', 연구자: '{researcher}'")

                if topic and researcher:
                    combinations.add((topic, researcher))
                    logging.info(f"조합 추가: ({topic}, {researcher})")
        else:
            logging.error("적요 컬럼이 연구비 데이터에 없습니다.")

        result = sorted(list(combinations))
        logging.info(f"최종 추출된 조합: {result}")
        return result

    def _filter_data_by_topic_researcher(self, research_data: pd.DataFrame, topic: str, researcher: str) -> pd.DataFrame:
        '''특정 연구주제와 연구자의 데이터만 필터링합니다.'''
        if '적요' not in research_data.columns:
            return pd.DataFrame()

        filtered_rows = []
        for _, row in research_data.iterrows():
            summary = row['적요']
            row_topic = self._extract_research_topic(summary)
            row_researcher = self._extract_researcher_name(summary)

            if row_topic == topic and row_researcher == researcher:
                filtered_rows.append(row)

        return pd.DataFrame(filtered_rows) if filtered_rows else pd.DataFrame()

    def _generate_individual_table(self, filtered_data: pd.DataFrame, topic: str, researcher: str) -> pd.DataFrame:
        '''개별 연구주제/연구자 표를 생성합니다.'''
        try:
            # 기본 구조 생성 (총액 행 제외)
            base_summary = self.generate_summary_sheet(filtered_data)

            if base_summary.empty:
                # 빈 기본 구조 생성
                base_summary = self._create_empty_budget_structure()

            # 총액 행 제거
            result_data = base_summary[base_summary['예산목'] != '총액'].copy()

            # 연구개발비와 유형자산 행 추가
            research_dev_expense = self._calculate_research_expense(filtered_data, '연구개발비')
            asset_expense = self._calculate_research_expense(filtered_data, '자산취득비')

            research_dev_row = {
                '예산목': '연구개발비',
                '세목': '연구개발비',
                '예산과목': '연구개발비',
                '예산금액': 0,
                '지출액': research_dev_expense,
                '예산잔액': -research_dev_expense,
                '집행률': '0%' if research_dev_expense == 0 else '∞%'
            }

            asset_row = {
                '예산목': '유형자산',
                '세목': '자산취득비',
                '예산과목': '자산취득비',
                '예산금액': 0,
                '지출액': asset_expense,
                '예산잔액': -asset_expense,
                '집행률': '0%' if asset_expense == 0 else '∞%'
            }

            result_data = pd.concat([
                result_data,
                pd.DataFrame([research_dev_row]),
                pd.DataFrame([asset_row])
            ], ignore_index=True)

            # 제목 행 추가 (주제와 연구자)
            title_row = {
                '예산목': topic,
                '세목': researcher,
                '예산과목': '',
                '예산금액': '',
                '지출액': '',
                '예산잔액': '',
                '집행률': ''
            }

            # 빈 행들 추가 (구분을 위해 2개)
            empty_rows = []
            for _ in range(2):
                empty_row = {col: '' for col in SUMMARY_SHEET_COLUMNS}
                empty_rows.append(empty_row)

            result_data = pd.concat([
                pd.DataFrame(empty_rows),
                pd.DataFrame([title_row]),
                result_data
            ], ignore_index=True)

            # 개별 표 뒤에도 2개의 빈 행 추가 (다음 표와의 구분용)
            end_empty_rows = []
            for _ in range(2):
                empty_row = {col: '' for col in SUMMARY_SHEET_COLUMNS}
                end_empty_rows.append(empty_row)

            result_data = pd.concat([result_data, pd.DataFrame(end_empty_rows)], ignore_index=True)

            return result_data

        except Exception as e:
            logging.error(f"개별 표 생성 중 오류: {str(e)}")
            return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

    def _create_empty_budget_structure(self) -> pd.DataFrame:
        '''빈 예산 구조를 생성합니다.'''
        try:
            # 기본 구조 생성
            hierarchical_data = self._create_hierarchical_structure(pd.DataFrame())

            if hierarchical_data.empty:
                # 수동으로 기본 구조 생성
                budget_categories = self.budget_classification['budget_categories']
                rows = []

                for budget_category, category_info in budget_categories.items():
                    is_first_category = True
                    for subcategory, budget_items in category_info['subcategories'].items():
                        is_first_subcategory = True
                        for budget_item in budget_items:
                            row = {
                                '예산목': budget_category if is_first_category else '',
                                '세목': subcategory if is_first_subcategory else '',
                                '예산과목': budget_item,
                                '예산금액': 0,
                                '지출액': 0,
                                '예산잔액': 0,
                                '집행률': '0%'
                            }
                            rows.append(row)
                            is_first_category = False
                            is_first_subcategory = False

                hierarchical_data = pd.DataFrame(rows)

            return hierarchical_data

        except Exception as e:
            logging.error(f"빈 예산 구조 생성 중 오류: {str(e)}")
            return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

    def _calculate_research_expense(self, research_data: pd.DataFrame, category: str) -> float:
        '''연구비 데이터에서 특정 카테고리의 지출액을 계산합니다.'''
        try:
            if research_data is None or research_data.empty:
                return 0.0

            if '총지급액' not in research_data.columns:
                logging.warning("연구비 데이터에 '총지급액' 컬럼이 없습니다.")
                return 0.0

            # 카테고리별 키워드 매핑
            category_keywords = {
                '연구개발비': ['연구개발비', '연구비', '연구개발'],
                '자산취득비': ['자산취득비', '자산취득', '유형자산', '장비구입', '기자재']
            }

            if category not in category_keywords:
                return 0.0

            keywords = category_keywords[category]
            total_expense = 0.0

            # 예산과목이나 적요에서 키워드를 포함하는 항목들의 지출액 합계
            for _, row in research_data.iterrows():
                budget_item = str(row.get('예산과목', ''))
                summary = str(row.get('적요', ''))
                expense = pd.to_numeric(row.get('총지급액', 0), errors='coerce')

                if pd.isna(expense):
                    expense = 0.0

                # 키워드 매칭 확인
                for keyword in keywords:
                    if keyword in budget_item or keyword in summary:
                        total_expense += expense
                        break  # 중복 계산 방지

            return total_expense

        except Exception as e:
            logging.error(f"연구비 지출액 계산 중 오류: {str(e)}")
            return 0.0

    def get_summary_data(self) -> Optional[pd.DataFrame]:
        '''생성된 요약 데이터를 반환합니다.'''
        return self.summary_data.copy() if self.summary_data is not None else None


class DashboardGenerator:
    '''대시보드 시트 생성 클래스 - 현대적 High-end Company 스타일'''

    def __init__(self):
        self.budget_classification = BUDGET_CLASSIFICATION

        # 현대적 High-end Company 색상 팔레트 (검정색 배경 버전)
        self.color_palette = {
            'primary_black': '1C1C1C',     # 메인 검정색 배경 (Excel 호환, 더 진한 회색)
            'dark_gray': '1F1F1F',         # 어두운 회색 (카드 배경)
            'silver_accent': 'C0C0C0',     # 실버 액센트 (제목, 강조)
            'white_text': 'FFFFFF',        # 화이트 텍스트
            'light_gray': 'E5E7EB',        # 밝은 회색 (부제목)
            'translucent_gray': '404040',  # 반투명 회색 (레이아웃 배경)
            'success_green': '10B981',     # 성공 지표 (좋은 집행률)
            'warning_orange': 'F59E0B',    # 경고 지표 (보통 집행률)
            'danger_red': 'EF4444',        # 위험 지표 (낮은 집행률)
            'info_blue': '3B82F6',         # 정보 지표 (센터)
            'purple_accent': '8B5CF6'      # 보라 액센트 (심층연구)
        }

    def generate_dashboard_sheet(self, total_sheet_data: pd.DataFrame) -> pd.DataFrame:
        '''
        총액 시트 데이터를 기반으로 대시보드 시트를 생성합니다.

        Args:
            total_sheet_data (pd.DataFrame): 총액 시트 데이터

        Returns:
            pd.DataFrame: 대시보드용 데이터 (빈 DataFrame - 실제 대시보드는 Excel에서 직접 생성)
        '''
        try:
            logging.info("2025 차세대 국가대표 스포츠과학지원 사업 예산 현황 대시보드 생성 시작 - 검정색 배경 적용")

            # 대시보드는 Excel 워크시트에서 직접 생성하므로 빈 DataFrame 반환
            # 실제 대시보드 생성은 create_dashboard_in_worksheet 메서드에서 수행
            return pd.DataFrame()

        except Exception as e:
            logging.error(f"대시보드 생성 중 오류: {str(e)}")
            return pd.DataFrame()

    def create_dashboard_in_worksheet(self, worksheet, total_sheet_data: pd.DataFrame):
        '''
        Excel 워크시트에 직접 대시보드를 생성합니다. (현대적 High-end Company 스타일)
        총액 시트를 Excel 수식으로 참조하여 실시간 연동됩니다.

        Args:
            worksheet: openpyxl 워크시트 객체
            total_sheet_data (pd.DataFrame): 총액 시트 데이터 (참조 위치 찾기용)
        '''
        try:
            logging.info("현대적 High-end Company 대시보드 생성 시작 - 총액 시트 실시간 연동")

            # 1. 총액 시트 참조 정보 찾기
            excel_refs = self._find_total_row_in_sheet(total_sheet_data)

            # 2. 대시보드 레이아웃 설정 (현대적 스타일)
            self._setup_modern_dashboard_layout(worksheet)

            # 3. 제목 및 헤더 생성 (현대적 테마)
            self._create_modern_dashboard_header(worksheet)

            # 4. 사업 기본 정보 섹션 생성 (B5에 배치)
            self._create_project_info_section(worksheet)

            # 5. KPI 지표 섹션 생성 (총액 시트 참조)
            self._create_modern_kpi_section(worksheet, excel_refs)

            # 6. 차트 섹션 생성 (총액 시트 참조)
            self._create_modern_chart_section(worksheet, excel_refs)

            # 7. 예산과목별 지표 섹션 생성 (B25에 추가) - 위치 조정
            self._create_budget_item_indicators_section(worksheet, total_sheet_data)

            # 8. 현대적 스타일링 적용
            self._apply_modern_dashboard_styling(worksheet)

            # 9. 최종 섹션 구분선 추가 - 위치 조정
            self._add_section_divider(worksheet, 'B48', 'K48', '대시보드 완료')

        except Exception as e:
            logging.error(f"대시보드 워크시트 생성 중 오류: {str(e)}")

    def _find_total_row_in_sheet(self, total_sheet_data: pd.DataFrame) -> dict:
        '''총액 시트에서 총액 행의 위치를 찾아 Excel 참조 정보를 반환합니다.'''
        try:
            logging.info("총액 시트에서 총액 행 위치 찾기 시작")

            if total_sheet_data.empty:
                return {
                    'total_row_index': None,
                    'budget_col': 'D',  # 예산금액 컬럼 (D)
                    'center_col': 'E',  # 센터 컬럼 (E)
                    'research_col': 'F',  # 심층연구 컬럼 (F)
                    'remaining_col': 'G',  # 예산잔액 컬럼 (G)
                    'execution_col': 'H'  # 집행률 컬럼 (H)
                }

            # 총액 행 찾기 (DataFrame의 인덱스 + 2 = Excel 행 번호, 헤더 때문에)
            total_row_index = None
            for idx, row in total_sheet_data.iterrows():
                if row['예산목'] == '총액':
                    total_row_index = idx + 2  # DataFrame 인덱스 + 헤더(1) + Excel 1-based(1)
                    break

            if total_row_index is None:
                # 총액 행이 없으면 마지막 행 + 1로 설정 (총액 행이 추가될 위치)
                total_row_index = len(total_sheet_data) + 2

            excel_refs = {
                'total_row_index': total_row_index,
                'budget_col': 'D',  # 예산금액 컬럼
                'center_col': 'E',  # 센터 컬럼
                'research_col': 'F',  # 심층연구 컬럼
                'remaining_col': 'G',  # 예산잔액 컬럼
                'execution_col': 'H'  # 집행률 컬럼
            }

            logging.info(f"총액 행 위치 찾기 완료 - 행: {total_row_index}")
            return excel_refs

        except Exception as e:
            logging.error(f"총액 행 위치 찾기 중 오류: {str(e)}")
            return {
                'total_row_index': None,
                'budget_col': 'D',
                'center_col': 'E',
                'research_col': 'F',
                'remaining_col': 'G',
                'execution_col': 'H'
            }

    def _setup_dark_dashboard_layout(self, worksheet):
        '''검정색 배경의 대시보드 레이아웃을 설정합니다.'''
        try:
            logging.info("검정색 배경 대시보드 레이아웃 설정 시작")

            # 컬럼 너비 설정 (더 넓게)
            worksheet.column_dimensions['A'].width = 3   # 여백
            worksheet.column_dimensions['B'].width = 25  # 라벨 (KPI 카드와 통일)
            worksheet.column_dimensions['C'].width = 20  # 값
            worksheet.column_dimensions['D'].width = 25  # KPI 카드와 통일
            worksheet.column_dimensions['E'].width = 25  # 차트 영역
            worksheet.column_dimensions['F'].width = 25  # 차트 영역 (KPI 카드와 통일)
            worksheet.column_dimensions['G'].width = 25  # 차트 영역
            worksheet.column_dimensions['H'].width = 25  # KPI 카드와 통일
            worksheet.column_dimensions['J'].width = 25  # KPI 카드와 통일

            # 행 높이 설정 (더 높게)
            for row in range(1, 35):
                worksheet.row_dimensions[row].height = 30

            logging.info("검정색 배경 대시보드 레이아웃 설정 완료")

        except Exception as e:
            logging.error(f"레이아웃 설정 중 오류: {str(e)}")

    def _create_dark_dashboard_header(self, worksheet):
        '''검정색 배경의 대시보드 헤더를 생성합니다.'''
        try:
            from openpyxl.styles import Font, Alignment, PatternFill

            # 메인 제목 (흰색 텍스트)
            worksheet['B2'] = "2025 차세대 국가대표 스포츠과학지원 사업 예산 현황"
            worksheet['B2'].font = Font(name='맑은 고딕', size=28, bold=True, color='FFFFFF')  # 흰색
            worksheet['B2'].alignment = Alignment(horizontal='left', vertical='center')

            # 부제목 (밝은 회색 텍스트)
            worksheet['B3'] = "실시간 예산 집행 현황 및 KPI 지표 - Dark Theme"
            worksheet['B3'].font = Font(name='맑은 고딕', size=14, color='E0E0E0')  # 밝은 회색
            worksheet['B3'].alignment = Alignment(horizontal='left', vertical='center')

            # 구분선 (어두운 회색)
            worksheet['B4'].fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
            worksheet.merge_cells('B4:G4')

            logging.info("검정색 배경 대시보드 헤더 생성 완료")

        except Exception as e:
            logging.error(f"헤더 생성 중 오류: {str(e)}")

    def _setup_modern_dashboard_layout(self, worksheet):
        '''현대적 High-end Company 스타일의 대시보드 레이아웃을 설정합니다.'''
        try:
            from openpyxl.styles import PatternFill

            # 메인 검정색 배경
            black_fill = PatternFill(start_color=self.color_palette['primary_black'],
                                   end_color=self.color_palette['primary_black'], fill_type='solid')

            # 전체 워크시트 배경을 검정색으로 설정 (V열까지 확장)
            for row in range(1, 60):
                for col in range(1, 23):  # V열은 22번째 열이므로 23까지
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = black_fill

            # 고급스러운 컬럼 너비 설정 (V열까지 확장)
            column_widths = {
                'A': 3,   # 여백
                'B': 25,  # 메인 콘텐츠 (KPI 카드 폭 통일)
                'C': 20,  # 여백
                'D': 25,  # KPI 카드 (폭 확장)
                'E': 20,  # 여백
                'F': 25,  # KPI 카드 (폭 통일)
                'G': 20,  # 여백
                'H': 25,  # KPI 카드 (폭 통일)
                'I': 20,  # 여백
                'J': 25,  # KPI 카드 (폭 통일)
                'K': 18,  # 차트 영역
                'L': 18,  # 차트 영역
                'M': 18,  # 차트 영역
                'N': 18,  # 차트 영역
                'O': 18,  # 추가 영역
                'P': 18,  # 추가 영역
                'Q': 18,  # 추가 영역
                'R': 18,  # 추가 영역
                'S': 18,  # 추가 영역
                'T': 18,  # 추가 영역
                'U': 18,  # 추가 영역
                'V': 3    # 우측 여백
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # 고급스러운 행 높이 설정 (더 체계적이고 직관적인 간격)
            row_heights = {
                1: 25,   # 상단 여백 (더 넉넉하게)
                2: 45,   # 메인 제목 (더 크게)
                3: 30,   # 부제목 (더 여유있게)
                4: 10,   # 구분선 (더 두껍게)
                5: 20,   # 여백 (섹션 구분)
                6: 35,   # KPI 섹션 제목 (더 크게)
                7: 15,   # 여백
                8: 35,   # KPI 카드 제목 (더 크게)
                9: 35,   # KPI 카드 값 (더 크게)
                10: 8,   # 그림자 효과
                11: 25,  # 섹션 구분 여백
                12: 35,  # 확장 KPI 섹션 제목
                13: 40,  # 여백 (14번째와 동일하게)
                14: 40,  # 확장 KPI 카드 제목
                15: 45,  # 확장 KPI 카드 값
                16: 8,   # 그림자 효과
                17: 30,  # 차트 섹션 제목
                18: 20   # 차트 섹션 여백
            }

            for row, height in row_heights.items():
                worksheet.row_dimensions[row].height = height

            # 나머지 행들은 기본 높이
            for row in range(14, 60):
                worksheet.row_dimensions[row].height = 25

            logging.info("현대적 High-end Company 대시보드 레이아웃 설정 완료")

        except Exception as e:
            logging.error(f"대시보드 레이아웃 설정 중 오류: {str(e)}")

    def _create_modern_dashboard_header(self, worksheet):
        '''현대적 High-end Company 스타일의 대시보드 헤더를 생성합니다.'''
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # 메인 제목 (실버 텍스트 + 그림자 효과)
            worksheet['B2'] = "2025 차세대 국가대표 스포츠과학지원 사업 예산 현황"
            worksheet['B2'].font = Font(name='맑은 고딕', size=32, bold=True, color=self.color_palette['silver_accent'])
            worksheet['B2'].alignment = Alignment(horizontal='left', vertical='center')

            # 메인 제목 셀 병합 (더 넓게)
            worksheet.merge_cells('B2:H2')

            # 부제목 (밝은 회색 텍스트 + 이탤릭)
            worksheet['B3'] = "실시간 예산 집행 현황 및 KPI 지표 | Executive Dashboard"
            worksheet['B3'].font = Font(name='맑은 고딕', size=14, italic=True, color=self.color_palette['light_gray'])
            worksheet['B3'].alignment = Alignment(horizontal='left', vertical='center')

            # 부제목 셀 병합
            worksheet.merge_cells('B3:H3')

            # 고급스러운 구분선 (그라데이션 효과를 위한 여러 셀)
            silver_fill = PatternFill(start_color=self.color_palette['silver_accent'],
                                    end_color=self.color_palette['silver_accent'], fill_type='solid')

            # 구분선을 더 넓게 설정
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                worksheet[f'{col}4'].fill = silver_fill
                worksheet[f'{col}4'].border = Border(
                    top=Side(style='thin', color=self.color_palette['white_text']),
                    bottom=Side(style='thin', color=self.color_palette['white_text'])
                )

            worksheet.merge_cells('B4:K4')

            # 현재 날짜/시간 표시 (왼쪽 상단에 추가)
            from datetime import datetime
            current_time = datetime.now().strftime("%Y-%m-%d")
            worksheet['B1'] = f"업데이트: {current_time}"
            worksheet['B1'].font = Font(name='맑은 고딕', size=10, color=self.color_palette['light_gray'])
            worksheet['B1'].alignment = Alignment(horizontal='left', vertical='center')

        except Exception as e:
            logging.error(f"대시보드 헤더 생성 중 오류: {str(e)}")

    def _create_project_info_section(self, worksheet):
        '''사업 기본 정보 섹션을 생성합니다. (B5에 배치)'''
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from config import YEARLY_BUDGET_DATA
            
            # 사업 정보 배경 (반투명 회색)
            info_fill = PatternFill(start_color=self.color_palette['translucent_gray'],
                                  end_color=self.color_palette['translucent_gray'], fill_type='solid')
            
            # 고급스러운 테두리
            info_border = Border(
                left=Side(style='thin', color=self.color_palette['silver_accent']),
                right=Side(style='thin', color=self.color_palette['silver_accent']),
                top=Side(style='thin', color=self.color_palette['silver_accent']),
                bottom=Side(style='thin', color=self.color_palette['silver_accent'])
            )

            # 사업기간 정보 (B5:C5)
            worksheet['B5'] = "사업기간"
            worksheet['B5'].font = Font(name='맑은 고딕', size=11, bold=True, color=self.color_palette['white_text'])
            worksheet['B5'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet['B5'].fill = info_fill
            worksheet['B5'].border = info_border

            worksheet['C5'] = "2025.03.01 ~ 2026.02.28"
            worksheet['C5'].font = Font(name='맑은 고딕', size=11, color=self.color_palette['info_blue'])
            worksheet['C5'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet['C5'].fill = info_fill
            worksheet['C5'].border = info_border

            # 총예산 정보 (D5:E5)
            worksheet['D5'] = "총예산"
            worksheet['D5'].font = Font(name='맑은 고딕', size=11, bold=True, color=self.color_palette['white_text'])
            worksheet['D5'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet['D5'].fill = info_fill
            worksheet['D5'].border = info_border

            # 2025년 총예산 계산 (예산과목: 금액 딕셔너리의 합)
            total_budget = sum(YEARLY_BUDGET_DATA['2025'].values())
            worksheet['E5'] = f'=TEXT({total_budget},"#,##0")'
            worksheet['E5'].font = Font(name='맑은 고딕', size=11, color=self.color_palette['success_green'])
            worksheet['E5'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet['E5'].fill = info_fill
            worksheet['E5'].border = info_border

            # 사업진행률 정보 (F5:G5)
            worksheet['F5'] = "사업진행률"
            worksheet['F5'].font = Font(name='맑은 고딕', size=11, bold=True, color=self.color_palette['white_text'])
            worksheet['F5'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet['F5'].fill = info_fill
            worksheet['F5'].border = info_border

            # 사업진행률 계산 (현재 날짜 기준)
            from datetime import datetime, date
            start_date = date(2025, 3, 1)
            end_date = date(2026, 2, 28)
            current_date = datetime.now().date()
            # print(f"Current date: {current_date}, Start date: {start_date}, End date: {end_date}")
            
            if current_date < start_date:
                progress_rate = 0
            elif current_date > end_date:
                progress_rate = 100
            else:
                total_days = (end_date - start_date).days
                # print(f"Total days in project: {total_days}")
                elapsed_days = (current_date - start_date).days
                # print(f"Elapsed days since start: {elapsed_days}")
                progress_rate = (elapsed_days / total_days) * 100
                # print(f"Calculated progress rate: {progress_rate:.2f}%")

            worksheet['G5'] = f'"{progress_rate:.1f}%"'
            worksheet['G5'].font = Font(name='맑은 고딕', size=11, color=self.color_palette['warning_orange'])
            worksheet['G5'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet['G5'].fill = info_fill
            worksheet['G5'].border = info_border

            logging.info("사업 기본 정보 섹션 생성 완료")

        except Exception as e:
            logging.error(f"사업 기본 정보 섹션 생성 중 오류: {str(e)}")

    def _create_modern_kpi_section(self, worksheet, excel_refs: dict):
        '''현대적 스타일의 KPI 지표 섹션을 생성합니다. (총액 시트 참조)'''
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            logging.info("현대적 KPI 섹션 생성 시작 - 총액 시트 참조")

            # KPI 섹션 제목 (실버 텍스트) - 레이아웃에 맞춰 B6에 배치
            worksheet['B6'] = "핵심 성과 지표 (KPI)"
            worksheet['B6'].font = Font(name='맑은 고딕', size=18, bold=True, color=self.color_palette['silver_accent'])

            # KPI 카드 스타일 설정 (화이트 테두리)
            card_border = Border(
                left=Side(style='thin', color='FFFFFF'),
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='thin', color='FFFFFF'),
                bottom=Side(style='thin', color='FFFFFF')
            )

            # 총액 시트 참조가 있는 경우에만 수식 생성
            if excel_refs['total_row_index'] is not None:
                total_row = excel_refs['total_row_index']

                # 기존 KPI 카드들: 레이아웃(행 8/9)에 맞춰 배치
                # 총액 집행률 카드 (총액 시트 참조) - 동적 색상
                self._create_modern_kpi_card_with_formula(
                    worksheet, 'B8', '총액 집행률',
                    f'=ROUND((총액!{excel_refs["center_col"]}{total_row}+총액!{excel_refs["research_col"]}{total_row})/총액!{excel_refs["budget_col"]}{total_row}*100,1)&"%"',
                    self.color_palette['success_green'], card_border
                )

                # 인건비제외 집행률 카드 (인건비 항목 제외한 집행률)
                self._create_modern_kpi_card_with_formula(
                    worksheet, 'D8', '인건비제외 집행률',
                    f'=IFERROR(ROUND((총액!{excel_refs["center_col"]}{excel_refs["total_row_index"]}+총액!{excel_refs["research_col"]}{excel_refs["total_row_index"]})/(총액!{excel_refs["budget_col"]}{excel_refs["total_row_index"]}-총액!D4-총액!D5)*100,1)&"%","0%")',  # 인건비를 제외한 집행률 계산 (인건비 예산 제외)
                    self.color_palette['info_blue'], card_border
                )

                # 센터 집행률 카드 (총액 시트 참조)
                self._create_modern_kpi_card_with_formula(
                    worksheet, 'F8', '센터 집행률',
                    f'=ROUND(총액!{excel_refs["center_col"]}{total_row}/총액!{excel_refs["budget_col"]}{total_row}*100,1)&"%"',
                    self.color_palette['info_blue'], card_border
                )

                # 심층연구 집행률 카드 (총액 시트 참조)
                self._create_modern_kpi_card_with_formula(
                    worksheet, 'H8', '심층연구 집행률',
                    f'=ROUND(총액!{excel_refs["research_col"]}{total_row}/총액!{excel_refs["budget_col"]}{total_row}*100,1)&"%"',
                    self.color_palette['warning_orange'], card_border
                )

                # 예산 잔액 카드 (총액 시트 참조) - 천 단위 구분자 적용
                self._create_modern_kpi_card_with_formula(
                    worksheet, 'J8', '예산 잔액',
                    f'=TEXT(총액!{excel_refs["remaining_col"]}{total_row},"#,##0")',
                    self.color_palette['silver_accent'], card_border
                )

            else:
                # 총액 시트 참조가 없는 경우 기본값 (동일한 행 배치 규칙 적용)
                self._create_modern_kpi_card_with_formula(worksheet, 'B8', '총액 집행률', '0%', self.color_palette['success_green'], card_border)
                self._create_modern_kpi_card_with_formula(worksheet, 'D8', '인건비제외 집행률', '0%', self.color_palette['info_blue'], card_border)
                self._create_modern_kpi_card_with_formula(worksheet, 'F8', '센터 집행률', '0%', self.color_palette['info_blue'], card_border)
                self._create_modern_kpi_card_with_formula(worksheet, 'H8', '심층연구 집행률', '0%', self.color_palette['warning_orange'], card_border)
                self._create_modern_kpi_card_with_formula(worksheet, 'J8', '예산 잔액', '"0"', self.color_palette['silver_accent'], card_border)

        except Exception as e:
            logging.error(f"KPI 섹션 생성 중 오류: {str(e)}")



    def _create_modern_kpi_card_with_formula(self, worksheet, start_cell: str, title: str, formula: str, color: str, border):
        '''현대적 High-end Company 스타일의 개별 KPI 카드를 생성합니다. (Excel 수식 사용)'''
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # 고급스러운 카드 배경 (반투명 회색으로 고급스러움 추가)
            card_fill = PatternFill(start_color=self.color_palette['translucent_gray'],
                                  end_color=self.color_palette['translucent_gray'], fill_type='solid')

            # 고급스러운 테두리 (더 두껍고 세련된)
            premium_border = Border(
                left=Side(style='medium', color=self.color_palette['silver_accent']),
                right=Side(style='medium', color=self.color_palette['silver_accent']),
                top=Side(style='medium', color=self.color_palette['silver_accent']),
                bottom=Side(style='medium', color=self.color_palette['silver_accent'])
            )

            # 제목 셀 (화이트 텍스트 + 더 큰 폰트)
            worksheet[start_cell] = title
            worksheet[start_cell].font = Font(name='맑은 고딕', size=13, bold=True, color=self.color_palette['white_text'])
            worksheet[start_cell].alignment = Alignment(horizontal='center', vertical='center')
            worksheet[start_cell].fill = card_fill
            worksheet[start_cell].border = premium_border

            # 값 셀 (아래 행, 컬러 텍스트 + Excel 수식 + 더 큰 폰트)
            row = int(start_cell[1:]) + 1
            col = start_cell[0]
            value_cell = f"{col}{row}"

            worksheet[value_cell] = formula  # Excel 수식 입력
            worksheet[value_cell].font = Font(name='맑은 고딕', size=18, bold=True, color=color)
            worksheet[value_cell].alignment = Alignment(horizontal='center', vertical='center')
            worksheet[value_cell].fill = card_fill
            worksheet[value_cell].border = premium_border

            # 카드 하단에 미세한 그림자 효과 (다음 행에 어두운 선)
            shadow_row = row + 1
            shadow_cell = f"{col}{shadow_row}"
            shadow_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            worksheet[shadow_cell].fill = shadow_fill
            worksheet.row_dimensions[shadow_row].height = 3  # 얇은 그림자

            logging.info(f"고급 KPI 카드 생성 완료: {title} - {formula}")

        except Exception as e:
            logging.error(f"KPI 카드 생성 중 오류: {str(e)}")

    def _create_modern_chart_section(self, worksheet, excel_refs: dict):
        '''현대적 스타일의 차트 섹션을 생성합니다. (총액 시트 참조)'''
        try:
            from openpyxl.styles import Font

            logging.info("현대적 차트 섹션 생성 시작 - 총액 시트 참조")

            # 차트 섹션 제목 (확장 KPI 아래 시작)
            worksheet['B12'] = "데이터 시각화 차트"
            worksheet['B12'].font = Font(name='맑은 고딕', size=18, bold=True, color=self.color_palette['silver_accent'])

            # 집행률 소제목 (차트 섹션 하위)
            worksheet['B13'] = "집행률"
            worksheet['B13'].font = Font(name='맑은 고딕', size=16, bold=True, color=self.color_palette['silver_accent'])

            # 집행률 비교 차트 생성 (총액 시트 참조) - 차트는 14행부터 시작하도록 조정
            self._create_modern_execution_rate_chart(worksheet, excel_refs)

            # 예산 배분 섹션 제목 추가 (차트 섹션 하위)
            worksheet['B19'] = "예산 배분"
            worksheet['B19'].font = Font(name='맑은 고딕', size=16, bold=True, color=self.color_palette['silver_accent'])

            # 예산 vs 집행 현황 차트 생성 (총액 시트 참조)
            self._create_modern_budget_vs_execution_chart(worksheet, excel_refs)

        except Exception as e:
            logging.error(f"차트 섹션 생성 중 오류: {str(e)}")

    def _create_budget_item_indicators_section(self, worksheet, total_sheet_data: pd.DataFrame):
        '''대시보드에 예산과목별 지표 섹션을 생성합니다. (B25에 추가)'''
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            logging.info("대시보드 예산과목별 지표 섹션 생성 시작")

            # 검정색 테두리 스타일 정의
            dark_border = Border(
                left=Side(style='thin', color='000000'),    # 검정색
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # B25에 섹션 제목 추가 (대시보드 스타일에 맞춰) - 위치 조정
            worksheet['B25'] = "예산과목별 지표"
            worksheet['B25'].font = Font(name='맑은 고딕', size=16, bold=True, color=self.color_palette['silver_accent'])
            worksheet['B25'].alignment = Alignment(horizontal='left', vertical='center')

            # 차트 섹션과 동일한 스타일 적용 (반투명 회색 배경)
            chart_fill = PatternFill(start_color=self.color_palette['translucent_gray'],
                                   end_color=self.color_palette['translucent_gray'], fill_type='solid')

            # B26~H26에 헤더 추가 - 개선된 7개 컬럼 구조
            headers = ['예산목', '세목', '예산과목', '예산금액', '지출액', '예산잔액', '집행률(%)']
            header_cells = ['B26', 'C26', 'D26', 'E26', 'F26', 'G26', 'H26']

            for header, cell_ref in zip(headers, header_cells):
                cell = worksheet[cell_ref]
                cell.value = header
                cell.font = Font(name='맑은 고딕', size=12, bold=False, color=self.color_palette['white_text'])
                cell.alignment = Alignment(horizontal='center', vertical='center')  # 헤더 가운데 정렬
                cell.border = dark_border  # 검정색 테두리 적용
                # 헤더 배경은 차트 섹션과 동일한 반투명 회색으로 설정
                cell.fill = chart_fill

            # 예산과목별 데이터 추출 (총액 행 제외) - 개선된 7개 컬럼 구조
            budget_items_data = []
            for _, row in total_sheet_data.iterrows():
                if (row['예산과목'] and row['예산과목'] != '' and
                    row['예산목'] != '총액' and not pd.isna(row['예산과목'])):
                    budget_items_data.append({
                        '예산목': row['예산목'],
                        '세목': row['세목'],
                        '예산과목': row['예산과목'],
                        '예산금액': row['예산금액'],
                        '지출액': row['센터'] + row['심층연구'],  # 센터 + 심층연구 = 지출액
                        '예산잔액': row['예산잔액'],
                        '집행률': row['집행률']
                    })

            # 데이터 행 추가 (B27부터 시작) - 개선된 7개 컬럼 구조
            start_row = 27

            # 예산목과 세목 중복 처리를 위한 변수
            previous_budget_category = None
            previous_subcategory = None

            for i, item in enumerate(budget_items_data):
                row_num = start_row + i

                # 예산목 (B열) - 중복 시 빈 값으로 설정 (병합 전 준비)
                budget_category_cell = worksheet[f'B{row_num}']
                current_budget_category = item['예산목']
                if current_budget_category != previous_budget_category:
                    budget_category_cell.value = current_budget_category
                    previous_budget_category = current_budget_category
                else:
                    budget_category_cell.value = ""  # 중복된 경우 빈 값
                budget_category_cell.font = Font(name='맑은 고딕', size=10, color=self.color_palette['white_text'])
                budget_category_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                budget_category_cell.border = dark_border  # 검정색 테두리 적용
                budget_category_cell.fill = chart_fill  # 차트 섹션과 동일한 배경

                # 세목 (C열) - 중복 시 빈 값으로 설정 (병합 전 준비)
                subcategory_cell = worksheet[f'C{row_num}']
                current_subcategory = item['세목']
                if current_subcategory != previous_subcategory:
                    subcategory_cell.value = current_subcategory
                    previous_subcategory = current_subcategory
                else:
                    subcategory_cell.value = ""  # 중복된 경우 빈 값
                subcategory_cell.font = Font(name='맑은 고딕', size=10, color=self.color_palette['white_text'])
                subcategory_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                subcategory_cell.border = dark_border  # 검정색 테두리 적용
                subcategory_cell.fill = chart_fill  # 차트 섹션과 동일한 배경

                # 예산과목 (D열)
                budget_item_cell = worksheet[f'D{row_num}']
                budget_item_cell.value = item['예산과목']
                budget_item_cell.font = Font(name='맑은 고딕', size=10, color=self.color_palette['white_text'])
                budget_item_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                budget_item_cell.border = dark_border  # 검정색 테두리 적용
                budget_item_cell.fill = chart_fill  # 차트 섹션과 동일한 배경

                # 예산금액 (E열) - 총액 시트 참조
                budget_amount_cell = worksheet[f'E{row_num}']
                original_row = self._find_budget_item_row_in_total_sheet(total_sheet_data, item['예산과목'])
                if original_row:
                    budget_amount_cell.value = f'=총액!D{original_row}'  # 총액 시트의 D열(예산금액) 참조
                else:
                    budget_amount_cell.value = item['예산금액']
                budget_amount_cell.font = Font(name='맑은 고딕', size=10, color=self.color_palette['white_text'])
                budget_amount_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                budget_amount_cell.number_format = '#,##0'
                budget_amount_cell.border = dark_border  # 검정색 테두리 적용
                budget_amount_cell.fill = chart_fill  # 차트 섹션과 동일한 배경

                # 지출액 (F열) - 총액 시트의 센터+심층연구 참조
                expense_cell = worksheet[f'F{row_num}']
                if original_row:
                    expense_cell.value = f'=총액!E{original_row}+총액!F{original_row}'  # 총액 시트의 E열(센터)+F열(심층연구) 참조
                else:
                    expense_cell.value = item['지출액']
                expense_cell.font = Font(name='맑은 고딕', size=10, color=self.color_palette['white_text'])
                expense_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                expense_cell.number_format = '#,##0'
                expense_cell.border = dark_border  # 검정색 테두리 적용
                expense_cell.fill = chart_fill  # 차트 섹션과 동일한 배경

                # 예산잔액 (G열) - 총액 시트 참조
                remaining_cell = worksheet[f'G{row_num}']
                if original_row:
                    remaining_cell.value = f'=총액!G{original_row}'  # 총액 시트의 G열(예산잔액) 참조
                else:
                    remaining_cell.value = item['예산잔액']
                remaining_cell.font = Font(name='맑은 고딕', size=10, color=self.color_palette['white_text'])
                remaining_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                remaining_cell.number_format = '#,##0'
                remaining_cell.border = dark_border  # 검정색 테두리 적용
                remaining_cell.fill = chart_fill  # 차트 섹션과 동일한 배경

                # 집행률 (H열) - 총액 시트 참조
                execution_cell = worksheet[f'H{row_num}']
                if original_row:
                    execution_cell.value = f'=총액!H{original_row}'  # 총액 시트의 H열(집행률) 참조
                else:
                    execution_cell.value = f"{item['집행률']}"
                execution_cell.font = Font(name='맑은 고딕', size=10, color=self.color_palette['white_text'])
                execution_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                execution_cell.border = dark_border  # 검정색 테두리 적용
                execution_cell.fill = chart_fill  # 차트 섹션과 동일한 배경

            # 컬럼 너비 조정 (개선된 7개 컬럼 대시보드에 맞게)
            worksheet.column_dimensions['B'].width = 20  # 예산목
            worksheet.column_dimensions['C'].width = 20  # 세목
            worksheet.column_dimensions['D'].width = 20  # 예산과목 (KPI 카드와 통일)
            worksheet.column_dimensions['E'].width = 20  # 예산금액
            worksheet.column_dimensions['F'].width = 20  # 지출액
            worksheet.column_dimensions['G'].width = 20  # 예산잔액
            worksheet.column_dimensions['H'].width = 20  # 집행률
            worksheet.column_dimensions['I'].width = 20  # KPI 카드와 통일 (여백)
            worksheet.column_dimensions['J'].width = 23  # KPI 카드와 통일

            # 예산목 컬럼 병합 처리 (중복 제거)
            self._merge_budget_category_cells(worksheet, budget_items_data, start_row, chart_fill)

            # 세목 컬럼 병합 처리 (중복 제거)
            self._merge_subcategory_cells(worksheet, budget_items_data, start_row, chart_fill)

            # 예산과목별 지표 그래프 추가 (테이블 옆에) - 제거됨
            # self._create_budget_item_charts(worksheet, len(budget_items_data))

            logging.info(f"대시보드 예산과목별 지표 섹션 생성 완료: {len(budget_items_data)}개 항목")

        except Exception as e:
            logging.error(f"대시보드 예산과목별 지표 섹션 생성 중 오류: {str(e)}")

    def _merge_budget_category_cells(self, worksheet, budget_items_data: list, start_row: int, chart_fill):
        '''예산목 컬럼에서 중복되는 값들을 병합합니다.'''
        try:
            from openpyxl.styles import Font, Alignment, Border, Side

            logging.info("예산목 컬럼 병합 처리 시작")

            # 어두운 회색 테두리 스타일 정의
            dark_border = Border(
                left=Side(style='thin', color='404040'),
                right=Side(style='thin', color='404040'),
                top=Side(style='thin', color='404040'),
                bottom=Side(style='thin', color='404040')
            )

            # 예산목별 범위 계산
            merge_ranges = {}
            current_budget_category = None
            range_start = None

            for i, item in enumerate(budget_items_data):
                row_num = start_row + i
                budget_category = item['예산목']

                if budget_category != current_budget_category:
                    # 이전 범위 저장
                    if current_budget_category and range_start is not None:
                        merge_ranges[current_budget_category] = (range_start, row_num - 1)

                    # 새 범위 시작
                    current_budget_category = budget_category
                    range_start = row_num

            # 마지막 범위 저장
            if current_budget_category and range_start is not None:
                merge_ranges[current_budget_category] = (range_start, start_row + len(budget_items_data) - 1)

            # 병합 적용 (2개 이상의 행이 있는 경우에만)
            for budget_category, (start, end) in merge_ranges.items():
                if end > start:  # 2개 이상의 행이 있는 경우
                    merge_range = f'B{start}:B{end}'
                    worksheet.merge_cells(merge_range)
                    
                    # 병합된 셀에 텍스트와 스타일 적용
                    merged_cell = worksheet[f'B{start}']
                    merged_cell.value = budget_category
                    merged_cell.font = Font(name='맑은 고딕', size=10, color=self.color_palette['white_text'])
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                    merged_cell.border = dark_border  # 어두운 회색 테두리 적용
                    merged_cell.fill = chart_fill

                    logging.info(f"예산목 '{budget_category}' 병합 완료: {merge_range}")

            logging.info("예산목 컬럼 병합 처리 완료")

        except Exception as e:
            logging.error(f"예산목 컬럼 병합 중 오류: {str(e)}")

    def _merge_subcategory_cells(self, worksheet, budget_items_data: list, start_row: int, chart_fill):
        '''세목 컬럼에서 중복되는 값들을 병합합니다.'''
        try:
            from openpyxl.styles import Font, Alignment, Border, Side

            logging.info("세목 컬럼 병합 처리 시작")

            # 어두운 회색 테두리 스타일 정의
            dark_border = Border(
                left=Side(style='thin', color='404040'),
                right=Side(style='thin', color='404040'),
                top=Side(style='thin', color='404040'),
                bottom=Side(style='thin', color='404040')
            )

            # 세목별 범위 계산
            merge_ranges = {}
            current_subcategory = None
            range_start = None

            for i, item in enumerate(budget_items_data):
                row_num = start_row + i
                subcategory = item['세목']

                if subcategory != current_subcategory:
                    # 이전 범위 저장
                    if current_subcategory and range_start is not None:
                        merge_ranges[current_subcategory] = (range_start, row_num - 1)

                    # 새 범위 시작
                    current_subcategory = subcategory
                    range_start = row_num

            # 마지막 범위 저장
            if current_subcategory and range_start is not None:
                merge_ranges[current_subcategory] = (range_start, start_row + len(budget_items_data) - 1)

            # 병합 적용 (2개 이상의 행이 있는 경우에만)
            for subcategory, (start, end) in merge_ranges.items():
                if end > start:  # 2개 이상의 행이 있는 경우
                    merge_range = f'C{start}:C{end}'
                    worksheet.merge_cells(merge_range)
                    
                    # 병합된 셀에 텍스트와 스타일 적용
                    merged_cell = worksheet[f'C{start}']
                    merged_cell.value = subcategory
                    merged_cell.font = Font(name='맑은 고딕', size=10, color=self.color_palette['white_text'])
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                    merged_cell.border = dark_border  # 어두운 회색 테두리 적용
                    merged_cell.fill = chart_fill

                    logging.info(f"세목 '{subcategory}' 병합 완료: {merge_range}")

            logging.info("세목 컬럼 병합 처리 완료")

        except Exception as e:
            logging.error(f"세목 컬럼 병합 중 오류: {str(e)}")

    def _create_budget_item_charts(self, worksheet, data_count: int):
        '''예산과목별 지표 테이블 옆에 막대그래프를 생성합니다.'''
        try:
            from openpyxl.chart import BarChart, Reference

            logging.info("예산과목별 지표 그래프 생성 시작")

            # 1. 예산잔액 막대그래프 (E22:I30)
            budget_chart = BarChart()
            budget_chart.type = "col"
            budget_chart.style = 10
            budget_chart.title = None  # 제목 제거
            budget_chart.y_axis.title = None  # Y축 제목 제거
            budget_chart.x_axis.title = None  # X축 제목 제거

            # 예산잔액 데이터 참조 (G열) - 개선된 컬럼 구조에 맞게 업데이트
            budget_data = Reference(worksheet, min_col=7, min_row=26, max_row=26+data_count, max_col=7)  # G열: 예산잔액
            budget_categories = Reference(worksheet, min_col=4, min_row=27, max_row=26+data_count, max_col=4)  # D열: 예산과목

            budget_chart.add_data(budget_data, titles_from_data=False)
            budget_chart.set_categories(budget_categories)

            # 차트 스타일링
            budget_chart.width = 27
            budget_chart.height = 9
            budget_chart.legend = None  # 범례 제거

            # KPI와 동일한 색상 적용
            try:
                from openpyxl.drawing.fill import SolidColorFillProperties
                if len(budget_chart.series) > 0:
                    budget_chart.series[0].graphicalProperties.solidFill = SolidColorFillProperties(self.color_palette['primary_blue'])
            except Exception as color_e:
                logging.warning(f"예산잔액 차트 색상 적용 실패: {str(color_e)}")

            # 데이터 레이블 추가 (바깥쪽 끝에, 설명선과 항목이름 포함)
            try:
                from openpyxl.chart.label import DataLabelList
                if len(budget_chart.series) > 0:
                    budget_chart.series[0].dLbls = DataLabelList()
                    budget_chart.series[0].dLbls.showCatName = True  # 항목이름 표시
                    budget_chart.series[0].dLbls.showVal = True      # 값 표시
                    budget_chart.series[0].dLbls.showSerName = False # 계열 이름 제거
                    budget_chart.series[0].dLbls.showLegendKey = False # 범례 표지 제거
                    budget_chart.series[0].dLbls.showLeaderLines = True  # 설명선 표시
                    budget_chart.series[0].dLbls.position = 'outEnd'    # 바깥쪽 끝에 배치
            except Exception as label_e:
                logging.warning(f"예산잔액 차트 데이터 레이블 적용 실패: {str(label_e)}")

            # 차트를 워크시트에 추가 (위치 조정 - 새로운 테이블 너비에 맞게)
            worksheet.add_chart(budget_chart, "I26")

            # 2. 집행률 막대그래프 (E32:I40)
            execution_chart = BarChart()
            execution_chart.type = "col"
            execution_chart.style = 11
            execution_chart.title = None  # 제목 제거
            execution_chart.y_axis.title = None  # Y축 제목 제거
            execution_chart.x_axis.title = None  # X축 제목 제거

            # 집행률 데이터 참조 (H열) - 개선된 컬럼 구조에 맞게 업데이트
            execution_data = Reference(worksheet, min_col=8, min_row=26, max_row=26+data_count, max_col=8)  # H열: 집행률
            execution_categories = Reference(worksheet, min_col=4, min_row=27, max_row=26+data_count, max_col=4)  # D열: 예산과목

            execution_chart.add_data(execution_data, titles_from_data=False)
            execution_chart.set_categories(execution_categories)

            # 차트 스타일링
            execution_chart.width = 27
            execution_chart.height = 9
            execution_chart.legend = None  # 범례 제거

            # KPI와 동일한 색상 적용
            try:
                from openpyxl.drawing.fill import SolidColorFillProperties
                if len(execution_chart.series) > 0:
                    execution_chart.series[0].graphicalProperties.solidFill = SolidColorFillProperties(self.color_palette['secondary_green'])
            except Exception as color_e:
                logging.warning(f"집행률 차트 색상 적용 실패: {str(color_e)}")

            # 데이터 레이블 추가 (바깥쪽 끝에, 설명선과 항목이름 포함)
            try:
                from openpyxl.chart.label import DataLabelList
                if len(execution_chart.series) > 0:
                    execution_chart.series[0].dLbls = DataLabelList()
                    execution_chart.series[0].dLbls.showCatName = True  # 항목이름 표시
                    execution_chart.series[0].dLbls.showVal = True      # 값 표시
                    execution_chart.series[0].dLbls.showSerName = False # 계열 이름 제거
                    execution_chart.series[0].dLbls.showLegendKey = False # 범례 표지 제거
                    execution_chart.series[0].dLbls.showLeaderLines = True  # 설명선 표시
                    execution_chart.series[0].dLbls.position = 'outEnd'    # 바깥쪽 끝에 배치
            except Exception as label_e:
                logging.warning(f"집행률 차트 데이터 레이블 적용 실패: {str(label_e)}")

            # 차트를 워크시트에 추가 (위치 조정 - 새로운 테이블 너비에 맞게)
            worksheet.add_chart(execution_chart, "I35")

            logging.info("예산과목별 지표 그래프 생성 완료")

        except Exception as e:
            logging.error(f"예산과목별 지표 그래프 생성 중 오류: {str(e)}")

    def _find_budget_item_row_in_total_sheet(self, total_sheet_data: pd.DataFrame, budget_item: str) -> int:
        '''총액 시트에서 특정 예산과목의 행 번호를 찾습니다.'''
        try:
            for idx, row in total_sheet_data.iterrows():
                if row['예산과목'] == budget_item:
                    return idx + 2  # DataFrame 인덱스 + 헤더 행(1) + Excel 1-based(1)
            return None
        except Exception as e:
            logging.error(f"예산과목 행 찾기 중 오류: {str(e)}")
            return None

    def _create_modern_execution_rate_chart(self, worksheet, excel_refs: dict):
        '''현대적 스타일의 집행률 비교 차트를 생성합니다. (총액 시트 참조)'''
        try:
            from openpyxl.chart import BarChart, Reference
            from openpyxl.styles import Font, PatternFill, Border, Side

            # 차트 데이터 준비 (반투명 회색 배경으로 고급스러움 추가)
            chart_fill = PatternFill(start_color=self.color_palette['translucent_gray'],
                                   end_color=self.color_palette['translucent_gray'], fill_type='solid')
            
            # 테두리 스타일 정의 (검정색)
            dark_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )            # 집행률 차트 제목 추가 (D13에 차트 제목) - 위치 조정
            worksheet['D13'] = "집행률 비교"
            worksheet['D13'].font = Font(name='맑은 고딕', size=16, bold=True, color=self.color_palette['silver_accent'])

            # 위치 조정 (집행률 제목 아래 B14부터 시작)
            worksheet['B14'] = "구분"
            worksheet['C14'] = "집행률(%)"
            worksheet['B14'].font = Font(name='맑은 고딕', size=12, bold=False, color=self.color_palette['white_text'])
            worksheet['C14'].font = Font(name='맑은 고딕', size=12, bold=False, color=self.color_palette['white_text'])
            worksheet['B14'].alignment = Alignment(horizontal='center', vertical='center')  # 헤더 가운데 정렬
            worksheet['C14'].alignment = Alignment(horizontal='center', vertical='center')  # 헤더 가운데 정렬
            worksheet['B14'].fill = chart_fill
            worksheet['C14'].fill = chart_fill
            worksheet['B14'].border = dark_border  # 검정색 테두리 적용
            worksheet['C14'].border = dark_border  # 검정색 테두리 적용

            # 총액 시트 참조 수식으로 데이터 설정 (위치 조정)
            if excel_refs['total_row_index'] is not None:
                total_row = excel_refs['total_row_index']

                worksheet['B15'] = "총액"
                worksheet['C15'] = f'=ROUND((총액!{excel_refs["center_col"]}{total_row}+총액!{excel_refs["research_col"]}{total_row})/총액!{excel_refs["budget_col"]}{total_row}*100,1)'

                worksheet['B16'] = "센터"
                worksheet['C16'] = f'=ROUND(총액!{excel_refs["center_col"]}{total_row}/총액!{excel_refs["budget_col"]}{total_row}*100,1)'

                worksheet['B17'] = "심층연구"
                worksheet['C17'] = f'=ROUND(총액!{excel_refs["research_col"]}{total_row}/총액!{excel_refs["budget_col"]}{total_row}*100,1)'
            else:
                # 기본값
                worksheet['B15'] = "총액"
                worksheet['C15'] = 0
                worksheet['B16'] = "센터"
                worksheet['C16'] = 0
                worksheet['B17'] = "심층연구"
                worksheet['C17'] = 0

            # 데이터 셀 스타일링 (위치 조정)
            for row in range(15, 18):
                worksheet[f'B{row}'].font = Font(name='맑은 고딕', size=11, color=self.color_palette['white_text'])
                worksheet[f'C{row}'].font = Font(name='맑은 고딕', size=11, color=self.color_palette['white_text'])
                worksheet[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                worksheet[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                worksheet[f'B{row}'].fill = chart_fill
                worksheet[f'C{row}'].fill = chart_fill
                worksheet[f'B{row}'].border = dark_border  # 검정색 테두리 적용
                worksheet[f'C{row}'].border = dark_border  # 검정색 테두리 적용

            # 막대 차트 생성 (제목과 축 이름 제거) # TODO: y축이 100 (100%)인 곳에 빨간색 선 추가
            chart = BarChart()
            chart.type = "col"
            # chart.style = 2
            chart.title = None  # 차트 내부 제목 제거
            chart.y_axis.title = None  # Y축 제목 제거
            chart.x_axis.title = None  # X축 제목 제거

            # 데이터 범위 설정 (위치 조정 - B14부터 시작)
            data = Reference(worksheet, min_col=3, min_row=14, max_row=17, max_col=3)
            cats = Reference(worksheet, min_col=2, min_row=15, max_row=17)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # KPI와 동일한 색상 적용
            try:
                from openpyxl.drawing.fill import SolidColorFillProperties
                from openpyxl.drawing.colors import ColorChoice

                # 각 데이터 시리즈에 KPI 색상 적용
                if len(chart.series) > 0:
                    series = chart.series[0]
                    # 총액 - 녹색, 센터 - 파랑, 심층연구 - 주황
                    colors = [self.color_palette['success_green'],
                             self.color_palette['info_blue'],
                             self.color_palette['warning_orange']]

                    for i, color in enumerate(colors):
                        if i < len(series.dPt):
                            pt = series.dPt[i]
                            pt.spPr.solidFill = SolidColorFillProperties(srgbClr=color)
            except Exception:
                print(f"차트 색상 적용 실패{e}")
                # 색상 설정에 실패하면 기본 차트 사용
                pass

            # 데이터 라벨 추가 (숫자만 표시, 가운데 위치) - 폰트 크기 2포인트 증가
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False
            chart.dataLabels.showSerName = False
            chart.dataLabels.showLegendKey = False  # 범례 표지 제거
            chart.dataLabels.position = "ctr"
            chart.dataLabels.font = Font(name='맑은 고딕', size=13)

            # BUG: 차트 배경 색상 적용 안됨.
            # 차트 배경색 설정 (대시보드와 조화로운 어두운 회색)
            try:
                from openpyxl.drawing.fill import SolidColorFillProperties
                from openpyxl.drawing.colors import RgbColor

                # 차트 배경색 설정
                chart_bg_color = self.color_palette['translucent_gray']  # 반투명 회색

                # SolidColorFillProperties 올바른 사용법
                solid_fill = SolidColorFillProperties()
                solid_fill.srgbClr = RgbColor(chart_bg_color)

                # 차트 전체 배경 설정
                if hasattr(chart, 'chartSpace') and hasattr(chart.chartSpace, 'spPr'):
                    chart.chartSpace.spPr.solidFill = solid_fill

                # 플롯 영역 배경도 동일하게 설정
                if hasattr(chart, 'plotArea') and hasattr(chart.plotArea, 'spPr'):
                    plot_fill = SolidColorFillProperties()
                    plot_fill.srgbClr = RgbColor(chart_bg_color)
                    chart.plotArea.spPr.solidFill = plot_fill

            except Exception as e:
                logging.warning(f"차트 배경색 설정 실패: {str(e)}")

            # 차트 위치 설정 (좌측 배치 - KPI 영역 아래로 이동)
            chart.anchor = "D14"
            chart.width = 14
            chart.height = 10

            worksheet.add_chart(chart)

        except Exception as e:
            logging.error(f"집행률 차트 생성 중 오류: {str(e)}")

    def _create_modern_budget_vs_execution_chart(self, worksheet, excel_refs: dict):
        '''현대적 스타일의 예산 vs 집행 현황 차트를 생성합니다. (총액 시트 참조)'''
        try:
            from openpyxl.chart import PieChart, Reference
            from openpyxl.styles import Font, PatternFill, Border, Side

            # 차트 데이터 준비 (반투명 회색 배경으로 고급스러움 추가)
            chart_fill = PatternFill(start_color=self.color_palette['translucent_gray'],
                                   end_color=self.color_palette['translucent_gray'], fill_type='solid')
            
            # 테두리 스타일 정의 (검정색)
            dark_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # 예산 배분 차트 제목 추가 (H13에 차트 제목) - 위치 조정
            worksheet['H13'] = "예산 배분 현황"
            worksheet['H13'].font = Font(name='맑은 고딕', size=16, bold=True, color=self.color_palette['silver_accent'])

            # 예산 배분 차트 데이터 (예산 배분 제목 아래 B20부터 시작) - 위치 조정
            worksheet['B20'] = "구분"
            worksheet['C20'] = "금액"
            worksheet['B20'].font = Font(name='맑은 고딕', size=12, bold=True, color=self.color_palette['white_text'])
            worksheet['C20'].font = Font(name='맑은 고딕', size=12, bold=True, color=self.color_palette['white_text'])
            worksheet['B20'].alignment = Alignment(horizontal='center', vertical='center')  # 헤더 가운데 정렬
            worksheet['C20'].alignment = Alignment(horizontal='center', vertical='center')  # 헤더 가운데 정렬
            worksheet['B20'].fill = chart_fill
            worksheet['C20'].fill = chart_fill
            worksheet['B20'].border = dark_border  # 검정색 테두리 적용
            worksheet['C20'].border = dark_border  # 검정색 테두리 적용

            # 총액 시트 참조 수식으로 데이터 설정 (B21부터 시작) - 위치 조정, 천 단위 구분자 적용
            if excel_refs['total_row_index'] is not None:
                total_row = excel_refs['total_row_index']

                worksheet['B21'] = "센터"
                worksheet['C21'] = f'=총액!{excel_refs["center_col"]}{total_row}'

                worksheet['B22'] = "심층연구"
                worksheet['C22'] = f'=총액!{excel_refs["research_col"]}{total_row}'

                worksheet['B23'] = "예산잔액"
                worksheet['C23'] = f'=총액!{excel_refs["remaining_col"]}{total_row}'
            else:
                # 기본값
                worksheet['B21'] = "센터"
                worksheet['C21'] = 0
                worksheet['B22'] = "심층연구"
                worksheet['C22'] = 0
                worksheet['B23'] = "예산잔액"
                worksheet['C23'] = 0

            # 데이터 셀 스타일링 (예산 배분 차트) - 천 단위 구분자 적용, 위치 조정
            for row in range(20, 24):
                worksheet[f'B{row}'].font = Font(name='맑은 고딕', size=11, color=self.color_palette['white_text'])
                worksheet[f'C{row}'].font = Font(name='맑은 고딕', size=11, color=self.color_palette['white_text'])
                worksheet[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                worksheet[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
                worksheet[f'B{row}'].fill = chart_fill
                worksheet[f'C{row}'].fill = chart_fill
                worksheet[f'B{row}'].border = dark_border  # 검정색 테두리 적용
                worksheet[f'C{row}'].border = dark_border  # 검정색 테두리 적용
                # 금액 컬럼에 천 단위 구분자 적용
                if row >= 21:  # 데이터 행에만 적용 (헤더 제외)
                    worksheet[f'C{row}'].number_format = '#,##0'

            # 파이 차트 생성 (내부 제목 제거)
            chart = PieChart()
            chart.title = None  # 차트 내부 제목 제거
            # chart.style = 2

            # 데이터 범위 설정 (예산 배분 제목 아래 B20-C23 영역) - 위치 조정
            data = Reference(worksheet, min_col=3, min_row=20, max_row=23, max_col=3)
            cats = Reference(worksheet, min_col=2, min_row=21, max_row=23)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # KPI와 동일한 색상 적용 (파이 차트)
            try:
                from openpyxl.drawing.fill import SolidColorFillProperties
                from openpyxl.drawing.colors import ColorChoice

                # 각 데이터 포인트에 KPI 색상 적용
                if len(chart.series) > 0:
                    series = chart.series[0]
                    # 센터 - 파랑, 심층연구 - 주황, 예산잔액 - 실버
                    colors = [self.color_palette['info_blue'],      # 센터
                             self.color_palette['warning_orange'],   # 심층연구
                             self.color_palette['silver_accent']]   # 예산잔액

                    for i, color in enumerate(colors):
                        if i < len(series.dPt):
                            pt = series.dPt[i]
                            pt.spPr.solidFill = SolidColorFillProperties(srgbClr=color)
            except Exception:
                # 색상 설정에 실패하면 기본 차트 사용
                pass

            # 데이터 라벨 추가 (숫자만 표시, 가운데 위치) - 폰트 크기 2포인트 증가
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False
            chart.dataLabels.showSerName = False
            chart.dataLabels.showLegendKey = False  # 범례 표지 제거
            chart.dataLabels.position = "ctr"
            chart.dataLabels.font = Font(name='맑은 고딕', size=13)

            # 차트 배경색 설정 (대시보드와 조화로운 어두운 회색)
            try:
                from openpyxl.drawing.fill import SolidColorFillProperties
                from openpyxl.drawing.colors import RgbColor

                # 차트 배경색 설정
                chart_bg_color = self.color_palette['translucent_gray']  # 반투명 회색

                # SolidColorFillProperties 올바른 사용법
                solid_fill = SolidColorFillProperties()
                solid_fill.srgbClr = RgbColor(chart_bg_color)

                # 차트 전체 배경 설정
                if hasattr(chart, 'chartSpace') and hasattr(chart.chartSpace, 'spPr'):
                    chart.chartSpace.spPr.solidFill = solid_fill

                # 플롯 영역 배경도 동일하게 설정
                if hasattr(chart, 'plotArea') and hasattr(chart.plotArea, 'spPr'):
                    plot_fill = SolidColorFillProperties()
                    plot_fill.srgbClr = RgbColor(chart_bg_color)
                    chart.plotArea.spPr.solidFill = plot_fill

            except Exception as e:
                logging.warning(f"차트 배경색 설정 실패: {str(e)}")

            # 차트 위치 설정 (우측 배치 - KPI 영역 아래로 이동)
            chart.anchor = "H14"
            chart.width = 14
            chart.height = 10

            worksheet.add_chart(chart)

            logging.info("현대적 예산 vs 집행 차트 생성 완료")

        except Exception as e:
            logging.error(f"예산 vs 집행 차트 생성 중 오류: {str(e)}")

    def _apply_modern_dashboard_styling(self, worksheet):
        '''현대적 High-end Company 스타일의 대시보드 스타일링을 적용합니다.'''
        try:
            from openpyxl.styles import PatternFill

            logging.info("현대적 대시보드 스타일링 적용 시작")

            # 전체적인 검정색 배경 재확인
            black_fill = PatternFill(start_color=self.color_palette['primary_black'],
                                   end_color=self.color_palette['primary_black'], fill_type='solid')

            # 빈 셀들에 검정색 배경 적용 (V열까지 확장)
            for row in range(1, 50):
                for col in range(1, 23):  # V열은 22번째 열이므로 23까지
                    cell = worksheet.cell(row=row, column=col)
                    if cell.fill.start_color.index == '00000000':  # 기본 배경인 경우
                        cell.fill = black_fill

            logging.info("현대적 대시보드 스타일링 적용 완료")

        except Exception as e:
            logging.error(f"스타일링 적용 중 오류: {str(e)}")

    def _add_section_divider(self, worksheet, start_cell: str, end_cell: str, label: str = ""):
        '''섹션 구분선을 추가합니다. (사용자 경험 개선)'''
        try:
            from openpyxl.styles import PatternFill, Font, Alignment

            # 구분선 색상 (실버 그라데이션)
            divider_fill = PatternFill(start_color=self.color_palette['silver_accent'],
                                     end_color=self.color_palette['silver_accent'], fill_type='solid')

            # 구분선 생성
            start_col = start_cell[0]
            end_col = end_cell[0]
            row = start_cell[1:]

            # 구분선 셀들에 골드 배경 적용
            for col_ord in range(ord(start_col), ord(end_col) + 1):
                col = chr(col_ord)
                cell = f"{col}{row}"
                worksheet[cell].fill = divider_fill

                # 라벨이 있으면 중앙에 표시
                if label and col == chr((ord(start_col) + ord(end_col)) // 2):
                    worksheet[cell] = f"{label}"
                    worksheet[cell].font = Font(name='맑은 고딕', size=10, bold=True, color=self.color_palette['primary_black'])
                    worksheet[cell].alignment = Alignment(horizontal='center', vertical='center')

            # 구분선 병합
            worksheet.merge_cells(f'{start_cell}:{end_cell}')

        except Exception as e:
            logging.error(f"섹션 구분선 추가 중 오류: {str(e)}")

class TotalSheetGenerator:
    '''총액 시트 생성 클래스 (센터와 심층연구로 지출액 분리)'''

    def __init__(self):
        self.budget_classification = BUDGET_CLASSIFICATION

    def generate_total_sheet(self, business_data: pd.DataFrame,
                           research_data: pd.DataFrame) -> pd.DataFrame:
        '''
        사업비와 연구비 데이터를 통합하여 총액 시트를 생성합니다.
        사업비 표와 완전히 동일한 계층구조를 가집니다.

        Args:
            business_data (pd.DataFrame): 사업비 집행관리 데이터
            research_data (pd.DataFrame): 연구비 집행관리 데이터

        Returns:
            pd.DataFrame: 총액 시트 데이터
        '''
        try:
            logging.info("총액 시트 생성 시작")

            # 1. 사업비와 연구비 데이터 집계
            business_summary = self._aggregate_business_expenses(business_data)
            research_summary = self._aggregate_research_expenses(research_data)
            logging.info(f"데이터 집계 완료 - 사업비: {len(business_summary)}건, 연구비: {len(research_summary)}건")

            # 2. 사업비 표와 동일한 전체 계층구조 생성
            hierarchical_data = self._create_full_hierarchical_structure()
            logging.info(f"전체 계층구조 생성 완료: {len(hierarchical_data)}건")

            # 3. 각 예산과목에 센터/심층연구 지출액 매핑
            final_data = self._map_expenses_to_structure(hierarchical_data, business_summary, research_summary)

            # 4. 예산금액 및 계산 컬럼 추가
            final_data_with_calculations = self._add_budget_calculations(final_data)

            # 5. 총액 행 추가
            final_data_with_total = self._add_total_row(final_data_with_calculations)

            logging.info(f"총액 시트 생성 완료: {len(final_data_with_total)}건")
            return final_data_with_total

        except Exception as e:
            logging.error(f"총액 시트 생성 중 오류: {str(e)}")
            return pd.DataFrame(columns=TOTAL_SHEET_COLUMNS)

    def _aggregate_business_expenses(self, business_data: pd.DataFrame) -> pd.DataFrame:
        '''사업비 데이터를 예산과목별로 집계합니다.'''
        try:
            if business_data is None or business_data.empty:
                return pd.DataFrame(columns=['예산과목', '센터'])

            # 지출액 컬럼 확인 (총지급액 또는 지출액)
            amount_column = '총지급액' if '총지급액' in business_data.columns else '지출액'
            if amount_column not in business_data.columns:
                logging.error(f"사업비 데이터에 지출액 컬럼({amount_column})이 없습니다.")
                return pd.DataFrame(columns=['예산과목', '센터'])

            # 예산과목별 지출액 집계
            aggregated = business_data.groupby('예산과목')[amount_column].sum().reset_index()
            aggregated.columns = ['예산과목', '센터']

            logging.info(f"사업비 집계 결과: {len(aggregated)}개 예산과목")
            return aggregated

        except Exception as e:
            logging.error(f"사업비 집계 중 오류: {str(e)}")
            return pd.DataFrame(columns=['예산과목', '센터'])

    def _aggregate_research_expenses(self, research_data: pd.DataFrame) -> pd.DataFrame:
        '''연구비 데이터를 예산과목별로 집계합니다.'''
        try:
            if research_data is None or research_data.empty:
                return pd.DataFrame(columns=['예산과목', '심층연구'])

            # 지출액 컬럼 확인 (총지급액 또는 지출액)
            amount_column = '총지급액' if '총지급액' in research_data.columns else '지출액'
            if amount_column not in research_data.columns:
                logging.error(f"연구비 데이터에 지출액 컬럼({amount_column})이 없습니다.")
                return pd.DataFrame(columns=['예산과목', '심층연구'])

            # 예산과목별 지출액 집계
            aggregated = research_data.groupby('예산과목')[amount_column].sum().reset_index()
            aggregated.columns = ['예산과목', '심층연구']

            logging.info(f"연구비 집계 결과: {len(aggregated)}개 예산과목")
            return aggregated

        except Exception as e:
            logging.error(f"연구비 집계 중 오류: {str(e)}")
            return pd.DataFrame(columns=['예산과목', '심층연구'])

    def _merge_business_research_data(self, business_summary: pd.DataFrame,
                                    research_summary: pd.DataFrame) -> pd.DataFrame:
        '''사업비와 연구비 데이터를 예산과목 기준으로 통합합니다.'''
        try:
            # 모든 예산과목 목록 생성
            all_budget_items = set()
            if not business_summary.empty:
                all_budget_items.update(business_summary['예산과목'].tolist())
            if not research_summary.empty:
                all_budget_items.update(research_summary['예산과목'].tolist())

            # 기본 DataFrame 생성
            result_data = []
            for budget_item in sorted(all_budget_items):
                # 사업비 지출액 (센터)
                center_expense = 0
                if not business_summary.empty:
                    center_row = business_summary[business_summary['예산과목'] == budget_item]
                    if not center_row.empty:
                        center_expense = center_row['센터'].iloc[0]

                # 연구비 지출액 (심층연구)
                research_expense = 0
                if not research_summary.empty:
                    research_row = research_summary[research_summary['예산과목'] == budget_item]
                    if not research_row.empty:
                        research_expense = research_row['심층연구'].iloc[0]

                result_data.append({
                    '예산과목': budget_item,
                    '센터': center_expense,
                    '심층연구': research_expense
                })

            result_df = pd.DataFrame(result_data)
            logging.info(f"데이터 통합 완료: {len(result_df)}개 예산과목")
            return result_df

        except Exception as e:
            logging.error(f"데이터 통합 중 오류: {str(e)}")
            return pd.DataFrame(columns=['예산과목', '센터', '심층연구'])

    def _create_full_hierarchical_structure(self) -> pd.DataFrame:
        '''사업비 표와 완전히 동일한 전체 계층구조를 생성합니다.'''
        try:
            hierarchical_data = []

            # 예산 분류 구조를 순회하면서 모든 항목 생성
            for budget_category, category_data in self.budget_classification['budget_categories'].items():
                for subcategory, items in category_data['subcategories'].items():
                    for budget_item in items:
                        hierarchical_data.append({
                            '예산목': budget_category,
                            '세목': subcategory,
                            '예산과목': budget_item,
                            '센터': 0,  # 기본값 0으로 초기화
                            '심층연구': 0  # 기본값 0으로 초기화
                        })

            result_df = pd.DataFrame(hierarchical_data)
            logging.info(f"전체 계층구조 생성 완료: {len(result_df)}건")
            return result_df

        except Exception as e:
            logging.error(f"전체 계층구조 생성 중 오류: {str(e)}")
            return pd.DataFrame(columns=TOTAL_SHEET_COLUMNS)

    def _map_expenses_to_structure(self, hierarchical_data: pd.DataFrame,
                                 business_summary: pd.DataFrame,
                                 research_summary: pd.DataFrame) -> pd.DataFrame:
        '''전체 계층구조에 센터/심층연구 지출액을 매핑합니다.'''
        try:
            result_data = hierarchical_data.copy()

            # 사업비 지출액 매핑 (센터)
            if not business_summary.empty:
                for _, row in business_summary.iterrows():
                    budget_item = row['예산과목']
                    center_expense = row['센터']

                    # 해당 예산과목의 행을 찾아서 센터 지출액 업데이트
                    mask = result_data['예산과목'] == budget_item
                    if mask.any():
                        result_data.loc[mask, '센터'] = center_expense

            # 연구비 지출액 매핑 (심층연구)
            if not research_summary.empty:
                for _, row in research_summary.iterrows():
                    budget_item = row['예산과목']
                    research_expense = row['심층연구']

                    # 해당 예산과목의 행을 찾아서 심층연구 지출액 업데이트
                    mask = result_data['예산과목'] == budget_item
                    if mask.any():
                        result_data.loc[mask, '심층연구'] = research_expense

            logging.info("지출액 매핑 완료")
            return result_data

        except Exception as e:
            logging.error(f"지출액 매핑 중 오류: {str(e)}")
            return hierarchical_data

    def _create_hierarchical_structure(self, total_data: pd.DataFrame) -> pd.DataFrame:
        '''예산과목 데이터를 계층적 구조(예산목-세목-예산과목)로 변환합니다.'''
        try:
            if total_data.empty:
                return pd.DataFrame(columns=TOTAL_SHEET_COLUMNS)

            hierarchical_data = []

            for _, row in total_data.iterrows():
                budget_item = row['예산과목']
                center_expense = row['센터']
                research_expense = row['심층연구']

                # 예산 분류에서 해당 예산과목의 계층 정보 찾기
                budget_category, subcategory = self._find_budget_hierarchy(budget_item)

                hierarchical_data.append({
                    '예산목': budget_category,
                    '세목': subcategory,
                    '예산과목': budget_item,
                    '센터': center_expense,
                    '심층연구': research_expense
                })

            result_df = pd.DataFrame(hierarchical_data)
            logging.info(f"계층적 구조 변환 완료: {len(result_df)}건")
            return result_df

        except Exception as e:
            logging.error(f"계층적 구조 변환 중 오류: {str(e)}")
            return pd.DataFrame(columns=TOTAL_SHEET_COLUMNS)

    def _find_budget_hierarchy(self, budget_item: str) -> tuple:
        '''예산과목에서 예산목과 세목을 찾습니다.'''
        try:
            for budget_category, category_data in self.budget_classification['budget_categories'].items():
                for subcategory, items in category_data['subcategories'].items():
                    if budget_item in items:
                        return budget_category, subcategory

            # 찾지 못한 경우 기본값 반환
            logging.warning(f"예산과목 '{budget_item}'의 계층 정보를 찾을 수 없습니다.")
            return '기타', '기타'

        except Exception as e:
            logging.error(f"예산 계층 찾기 중 오류: {str(e)}")
            return '기타', '기타'

    def _add_budget_calculations(self, hierarchical_data: pd.DataFrame) -> pd.DataFrame:
        '''예산금액, 예산잔액, 집행률을 계산하여 추가합니다.'''
        try:
            if hierarchical_data.empty:
                return pd.DataFrame(columns=TOTAL_SHEET_COLUMNS)

            result_data = hierarchical_data.copy()
            default_budgets = self.budget_classification['2025_budget_amounts']

            # 예산금액 설정
            result_data['예산금액'] = result_data['예산과목'].apply(
                lambda x: self._get_budget_amount(x, default_budgets)
            )

            # 예산잔액 계산 (예산금액 - 센터 - 심층연구)
            result_data['예산잔액'] = (result_data['예산금액'] -
                                   result_data['센터'] -
                                   result_data['심층연구'])

            # 집행률 계산 ((센터 + 심층연구) / 예산금액 * 100)
            result_data['집행률'] = result_data.apply(
                lambda row: f"{((row['센터'] + row['심층연구']) / row['예산금액'] * 100):.0f}"
                if row['예산금액'] > 0 else "0", axis=1
            )

            # 컬럼 순서 정렬
            result_data = result_data[TOTAL_SHEET_COLUMNS]

            logging.info(f"예산 계산 완료: {len(result_data)}건")
            return result_data

        except Exception as e:
            logging.error(f"예산 계산 중 오류: {str(e)}")
            return pd.DataFrame(columns=TOTAL_SHEET_COLUMNS)

    def _get_budget_amount(self, budget_item: str, default_budgets: dict) -> int:
        '''예산과목의 예산금액을 반환합니다.'''
        return default_budgets.get(budget_item, 0)

    def _add_total_row(self, final_data: pd.DataFrame) -> pd.DataFrame:
        '''총액 행을 추가합니다.'''
        try:
            if final_data.empty:
                return final_data

            # 총액 계산
            total_budget = final_data['예산금액'].sum()
            total_center = final_data['센터'].sum()
            total_research = final_data['심층연구'].sum()
            total_remaining = total_budget - total_center - total_research
            total_execution_rate = f"{((total_center + total_research) / total_budget * 100):.0f}" if total_budget > 0 else "0"

            # 총액 행 생성
            total_row = {
                '예산목': '총액',
                '세목': '',
                '예산과목': '',
                '예산금액': total_budget,
                '센터': total_center,
                '심층연구': total_research,
                '예산잔액': total_remaining,
                '집행률': total_execution_rate
            }

            # 총액 행 추가
            result_data = pd.concat([final_data, pd.DataFrame([total_row])], ignore_index=True)

            logging.info("총액 행 추가 완료")
            return result_data

        except Exception as e:
            logging.error(f"총액 행 추가 중 오류: {str(e)}")
            return final_data


class InteractivePivotGenerator:
    '''xlwings를 사용한 대화형 피벗 테이블 생성 클래스'''

    def __init__(self):
        self.xlwings_available = self._check_xlwings_availability()

    def _check_xlwings_availability(self) -> bool:
        '''xlwings 사용 가능성 검사'''
        try:
            import xlwings as xw
            # Excel 애플리케이션 연결 테스트
            with xw.App(visible=False, add_book=False) as app:
                pass
            logging.info("xlwings 사용 가능 확인")
            return True
        except Exception as e:
            logging.warning(f"xlwings 사용 불가: {str(e)}")
            return False

    def add_interactive_features(self, file_path: str) -> bool:
        '''기존 Excel 파일에 대화형 피벗 테이블과 슬라이서 추가'''
        if not self.xlwings_available:
            logging.info("xlwings를 사용할 수 없어 대화형 기능을 건너뜁니다.")
            return False

        try:
            import xlwings as xw
            from config import PIVOT_CONFIG, PIVOT_SHEET_NAME, PIVOT_CHART_TITLE

            logging.info(f"대화형 피벗 테이블 생성 시작: {file_path}")

            with xw.App(visible=True) as app:
                wb = app.books.open(file_path)
                
                # 총액 시트가 존재하는지 확인
                if '총액' not in [sheet.name for sheet in wb.sheets]:
                    logging.error("총액 시트를 찾을 수 없습니다.")
                    return False

                # 피벗 테이블 시트 생성
                ws_pivot = wb.sheets.add(PIVOT_SHEET_NAME)
                
                # 피벗 테이블 생성
                pivot_table = self._create_pivot_table(wb, '총액', ws_pivot)
                if not pivot_table:
                    logging.error("피벗 테이블 생성 실패")
                    return False

                # 피벗 차트 추가
                self._add_pivot_chart(ws_pivot, pivot_table)

                # 슬라이서 추가
                self._add_slicers(wb, pivot_table, ws_pivot)

                # 파일 저장
                wb.save()
                logging.info("대화형 피벗 테이블 생성 완료")

            return True

        except Exception as e:
            logging.error(f"대화형 피벗 테이블 생성 중 오류: {str(e)}")
            return False

    def _create_pivot_data_sheet(self, wb, source_sheet_name: str) -> str:
        '''예산분석용 세로형 데이터 시트 생성'''
        try:
            import xlwings as xw
            
            logging.info("예산분석용 세로형 데이터 시트 생성 시작")
            
            # 원본 데이터 읽기
            ws_source = wb.sheets[source_sheet_name]
            source_range = ws_source.used_range
            
            logging.info(f"원본 데이터 범위: {source_range.address}, 행 수: {source_range.shape[0]}")
            
            # 새 시트 생성
            pivot_data_sheet_name = '예산분석데이터'
            try:
                # 기존 시트가 있으면 삭제
                wb.sheets[pivot_data_sheet_name].delete()
                logging.info("기존 예산분석데이터 시트 삭제")
            except:
                pass
            
            ws_pivot_data = wb.sheets.add(pivot_data_sheet_name)
            logging.info(f"새 예산분석데이터 시트 생성: {pivot_data_sheet_name}")
            
            # 헤더 설정
            ws_pivot_data.range('A1').value = '예산과목'
            ws_pivot_data.range('B1').value = '특성'
            ws_pivot_data.range('C1').value = '값'
            
            # 데이터 변환
            row_idx = 2  # 데이터 시작 행
            
            for data_row in range(2, source_range.shape[0] + 1):  # 헤더 제외
                budget_category = ws_source.range(f'C{data_row}').value  # 예산과목
                budget_amount = ws_source.range(f'D{data_row}').value    # 예산금액
                center_amount = ws_source.range(f'E{data_row}').value    # 센터
                research_amount = ws_source.range(f'F{data_row}').value  # 심층연구
                remaining_amount = ws_source.range(f'G{data_row}').value # 예산잔액
                execution_rate = ws_source.range(f'H{data_row}').value   # 집행률
                
                # 디버깅용 로그 추가
                if data_row <= 5:  # 처음 몇 행만 로깅
                    logging.debug(f"행 {data_row}: 예산과목={budget_category}, 예산금액={budget_amount}")
                
                # 예산과목이 비어있으면 건너뛰기
                if not budget_category or str(budget_category).strip() == '':
                    continue
                
                # None 값 처리 및 타입 변환
                try:
                    budget_amount = float(budget_amount) if budget_amount is not None else 0
                    center_amount = float(center_amount) if center_amount is not None else 0
                    research_amount = float(research_amount) if research_amount is not None else 0
                    remaining_amount = float(remaining_amount) if remaining_amount is not None else 0
                    execution_rate = float(execution_rate) if execution_rate is not None else 0
                except (ValueError, TypeError) as e:
                    logging.warning(f"행 {data_row} 숫자 형변환 실패: {e}, 0으로 설정")
                    budget_amount = center_amount = research_amount = remaining_amount = execution_rate = 0
                
                # 지출액 계산
                total_expense = center_amount + research_amount
                
                if budget_category:  # 예산과목이 있는 경우만
                    # 예산금액 행
                    ws_pivot_data.range(f'A{row_idx}').value = budget_category
                    ws_pivot_data.range(f'B{row_idx}').value = '예산금액'
                    ws_pivot_data.range(f'C{row_idx}').value = budget_amount
                    row_idx += 1
                    
                    # 지출액 행
                    ws_pivot_data.range(f'A{row_idx}').value = budget_category
                    ws_pivot_data.range(f'B{row_idx}').value = '지출액'
                    ws_pivot_data.range(f'C{row_idx}').value = total_expense
                    row_idx += 1
                    
                    # 예산잔액 행
                    ws_pivot_data.range(f'A{row_idx}').value = budget_category
                    ws_pivot_data.range(f'B{row_idx}').value = '예산잔액'
                    ws_pivot_data.range(f'C{row_idx}').value = remaining_amount
                    row_idx += 1
                    
                    # 집행률 행
                    ws_pivot_data.range(f'A{row_idx}').value = budget_category
                    ws_pivot_data.range(f'B{row_idx}').value = '집행률(%)'
                    ws_pivot_data.range(f'C{row_idx}').value = execution_rate
                    row_idx += 1
            
            logging.info(f"예산분석용 세로형 데이터 생성 완료: {row_idx-1}개 행")
            
            # 즉시 저장하여 데이터 확정
            try:
                wb.save()
                logging.info("예산분석 데이터 시트 저장 완료")
            except Exception as save_error:
                logging.warning(f"예산분석 데이터 시트 저장 실패: {save_error}")
            
            return pivot_data_sheet_name
            
        except Exception as e:
            logging.error(f"예산분석용 데이터 시트 생성 중 오류: {str(e)}")
            return None

    def _create_pivot_table(self, wb, source_sheet_name: str, ws_pivot) -> object:
        '''피벗 테이블 생성'''
        try:
            import xlwings as xw
            from config import PIVOT_CONFIG

            # 먼저 세로형 데이터 시트 생성
            pivot_data_sheet_name = self._create_pivot_data_sheet(wb, source_sheet_name)
            if not pivot_data_sheet_name:
                logging.error("피벗용 데이터 시트 생성 실패")
                return None

            # 세로형 데이터 시트를 소스로 사용
            ws_source = wb.sheets[pivot_data_sheet_name]
            
            # 세로형 데이터 범위 확인 (A:예산과목, B:특성, C:값)
            try:
                used_range = ws_source.api.UsedRange
                last_row = used_range.Row + used_range.Rows.Count - 1
                source_range = ws_source.range(f'A1:C{last_row}')
                logging.info(f"세로형 데이터 범위: {source_range.address}")
            except:
                # 수동으로 범위 찾기
                last_row = 1
                for row in range(1, 1000):  # 최대 1000행까지 확인
                    if ws_source.range(f'A{row}').value is not None:
                        last_row = row
                source_range = ws_source.range(f'A1:C{last_row}')
                logging.info(f"수동 범위 설정: {source_range.address}")

            logging.info(f"소스 데이터 범위: {source_range.address}")

            # 1. 피벗 캐시 생성
            pivot_cache = wb.api.PivotCaches().Create(
                SourceType=xw.constants.PivotTableSourceType.xlDatabase,
                SourceData=source_range.api
            )

            # 2. 피벗 테이블 생성
            logging.info("피벗 테이블 생성 중...")
            pivot_table = pivot_cache.CreatePivotTable(
                TableDestination=ws_pivot.range('A3').api,
                TableName='BudgetAnalysisPivot'
            )
            logging.info("피벗 테이블 기본 구조 생성 완료")

            # 3. 필드 배치 (새로운 세로형 구조: 예산과목, 특성, 값)
            logging.info("필드 배치 시작...")
            
            # 행 필드: 예산과목
            logging.info("예산과목 필드를 행 필드로 설정 중...")
            pivot_table.PivotFields('예산과목').Orientation = xw.constants.PivotFieldOrientation.xlRowField
            logging.info("예산과목 행 필드 설정 완료")

            # 열 필드: 특성 (예산금액, 지출액, 예산잔액, 집행률)
            logging.info("특성 필드를 열 필드로 설정 중...")
            pivot_table.PivotFields('특성').Orientation = xw.constants.PivotFieldOrientation.xlColumnField
            logging.info("특성 열 필드 설정 완료")

            # 값 필드: 값
            logging.info("값 필드를 데이터 필드로 설정 중...")
            try:
                # 사용 가능한 필드 확인
                available_fields = [field.Name for field in pivot_table.PivotFields()]
                logging.info(f"피벗 테이블 사용 가능한 필드들: {available_fields}")
                
                # 실제 값 필드명 찾기 (값, 값2, 값3 등 가능)
                value_field_name = None
                for field_name in available_fields:
                    if field_name.startswith('값'):
                        value_field_name = field_name
                        break
                
                if not value_field_name:
                    raise Exception("값 필드를 찾을 수 없습니다")
                
                logging.info(f"사용할 값 필드명: {value_field_name}")
                value_field = pivot_table.PivotFields(value_field_name)
                data_field = pivot_table.AddDataField(
                    value_field,
                    '값 합계',
                    xw.constants.ConsolidationFunction.xlSum
                )
                logging.info("값 데이터 필드 추가 완료")
                
                # 총합계 행과 열 제거
                logging.info("총합계 행과 열 제거 중...")
                try:
                    # 행 총합계 제거
                    pivot_table.RowGrand = False
                    logging.info("행 총합계 제거 완료")
                    
                    # 열 총합계 제거
                    pivot_table.ColumnGrand = False
                    logging.info("열 총합계 제거 완료")
                except Exception as grand_error:
                    logging.warning(f"총합계 제거 실패: {grand_error}")
                
                logging.info("피벗 테이블 필드 설정 완료")
                    
            except Exception as e:
                logging.error(f"데이터 필드 추가 중 오류: {str(e)}")
                return None

            logging.info("피벗 테이블 필드 구성 완료")
            return pivot_table

        except Exception as e:
            logging.error(f"피벗 테이블 생성 중 오류: {str(e)}")
            return None

    def _add_pivot_chart(self, ws_pivot, pivot_table):
        '''피벗 차트 추가'''
        try:
            import xlwings as xw
            from config import PIVOT_CONFIG, PIVOT_CHART_TITLE

            # 차트 위치 설정
            chart_pos = PIVOT_CONFIG['chart_position']

            # 피벗 차트 생성 (세로 막대형)
            chart_shape = ws_pivot.api.Shapes.AddChart2(
                227,  # 차트 스타일
                xw.constants.ChartType.xlColumnClustered
            )
            
            chart = chart_shape.Chart
            chart.SetSourceData(Source=ws_pivot.range('A3').expand().api)
            chart.HasTitle = True
            chart.ChartTitle.Text = PIVOT_CHART_TITLE

            # 차트 위치 조정
            chart_shape.Left = chart_pos['left']
            chart_shape.Top = chart_pos['top']
            chart_shape.Width = chart_pos['width']
            chart_shape.Height = chart_pos['height']

            logging.info("피벗 차트 추가 완료")

        except Exception as e:
            logging.error(f"피벗 차트 추가 중 오류: {str(e)}")

    def _add_slicers(self, wb, pivot_table, ws_pivot):
        '''슬라이서 추가'''
        try:
            import xlwings as xw
            from config import PIVOT_CONFIG

            slicer_positions = PIVOT_CONFIG['slicer_positions']

            # 1. 예산과목 슬라이서
            try:
                slicer_cache_budget = wb.api.SlicerCaches.Add2(
                    pivot_table,
                    '예산과목'
                )
                
                budget_pos = slicer_positions['budget_item']
                slicer_cache_budget.Slicers.Add(
                    SlicerDestination=ws_pivot.api,
                    Name='BudgetItemSlicer',
                    Caption='예산과목 선택',
                    Top=budget_pos['top'],
                    Left=budget_pos['left'],
                    Width=budget_pos['width'],
                    Height=budget_pos['height']
                )
                logging.info("예산과목 슬라이서 추가 완료")
                
            except Exception as e:
                logging.warning(f"예산과목 슬라이서 추가 실패: {str(e)}")

            # 2. 특성 필드 슬라이서 (예산금액, 지출액, 예산잔액, 집행률)
            try:
                logging.info("특성 필드 슬라이서 추가 중...")
                
                slicer_cache_characteristics = wb.api.SlicerCaches.Add2(
                    pivot_table,
                    '특성'
                )
                
                metric_pos = slicer_positions['metric']
                slicer_cache_characteristics.Slicers.Add(
                    SlicerDestination=ws_pivot.api,
                    Name='CharacteristicsSlicer',
                    Caption='측정항목 선택',
                    Top=metric_pos['top'],
                    Left=metric_pos['left'],
                    Width=metric_pos['width'],
                    Height=metric_pos['height']
                )
                logging.info("특성 필드 슬라이서 추가 완료")
                
            except Exception as e:
                logging.warning(f"특성 필드 슬라이서 추가 실패: {str(e)}")
                
        except Exception as e:
            logging.error(f"슬라이서 추가 중 오류: {str(e)}")

    def create_yearly_budget_comparison(self, file_path: str) -> bool:
        '''연도별 예산 비교용 피벗 테이블과 차트를 생성합니다.'''
        if not self.xlwings_available:
            logging.info("xlwings를 사용할 수 없어 연도별 비교 기능을 건너뜁니다.")
            return False

        try:
            import xlwings as xw
            from config import YEARLY_BUDGET_DATA, create_yearly_pivot_data, get_all_budget_items

            logging.info(f"연도별 예산 비교 테이블 생성 시작: {file_path}")

            with xw.App(visible=True, add_book=False) as app:
                wb = app.books.open(file_path)
                
                # 연도별 예산 데이터 시트 생성
                yearly_data_sheet_name = self._create_yearly_budget_data_sheet(wb)
                if not yearly_data_sheet_name:
                    logging.error("연도별 예산 데이터 시트 생성 실패")
                    return False

                # 연도별 비교 피벗 테이블 시트 생성 (기존 방식 유지)
                yearly_pivot_sheet_name = '연도별예산비교'
                try:
                    wb.sheets[yearly_pivot_sheet_name].delete()
                except:
                    pass
                
                ws_yearly_pivot = wb.sheets.add(yearly_pivot_sheet_name)
                
                # 대시보드 시트 확인 및 생성
                try:
                    ws_dashboard = wb.sheets['대시보드']
                    logging.info("기존 대시보드 시트 사용")
                except:
                    ws_dashboard = wb.sheets.add('대시보드')
                    logging.info("새 대시보드 시트 생성")
                
                # 연도별 피벗 테이블 생성 (연도별예산비교 시트에 생성, 대시보드에는 차트와 슬라이서만)
                yearly_pivot_table = self._create_yearly_pivot_table(wb, yearly_data_sheet_name, ws_yearly_pivot)
                if yearly_pivot_table:
                    # 대시보드에 제목 추가
                    self._add_yearly_comparison_title_to_dashboard(ws_dashboard)
                    
                    # 연도별 비교 차트를 대시보드 B53에 생성
                    self._add_yearly_comparison_chart_to_dashboard(wb, ws_dashboard, yearly_pivot_table)
                    
                    # 연도별 슬라이서를 대시보드 차트 아래에 추가
                    self._add_yearly_slicers_to_dashboard(wb, yearly_pivot_table, ws_dashboard)

                wb.save()
                logging.info("연도별 예산 비교 테이블 생성 완료 (대시보드 시트에 배치)")
                return True

        except Exception as e:
            logging.error(f"연도별 예산 비교 테이블 생성 중 오류: {str(e)}")
            return False

    def _add_yearly_comparison_title_to_dashboard(self, ws_dashboard):
        '''대시보드 시트에 연도별 예산 비교 섹션 제목을 추가합니다.'''
        try:
            import xlwings as xw
            
            # 제목 추가 (B50)
            title_cell = ws_dashboard.range('B50')
            title_cell.value = '📊 연도별 예산 비교 분석'
            
            # 제목 스타일 설정
            title_cell.api.Font.Size = 16
            title_cell.api.Font.Bold = True
            title_cell.api.Font.Color = 0x2E75B6  # 파란색
            
            # 부제목 추가 (B51)
            subtitle_cell = ws_dashboard.range('B51')
            subtitle_cell.value = '연도별 예산과목별 예산금액을 비교분석할 수 있습니다.'
            subtitle_cell.api.Font.Size = 11
            subtitle_cell.api.Font.Color = 0x595959  # 회색
            
            # 안내 메시지 추가 (B52)
            guide_cell = ws_dashboard.range('B52')
            guide_cell.value = '※ 우측 슬라이서를 사용하여 특정 예산과목이나 연도를 필터링할 수 있습니다.'
            guide_cell.api.Font.Size = 9
            guide_cell.api.Font.Color = 0x808080  # 연회색
            guide_cell.api.Font.Italic = True
            
            logging.info("대시보드 시트에 연도별 예산 비교 제목 및 안내문 추가 완료")
            
        except Exception as e:
            logging.warning(f"대시보드 시트 제목 추가 중 오류: {str(e)}")

    def _create_yearly_budget_data_sheet(self, wb) -> str:
        '''연도별 예산 데이터를 세로형으로 변환한 시트를 생성합니다.'''
        try:
            import xlwings as xw
            from config import YEARLY_BUDGET_DATA

            logging.info("연도별 예산 세로형 데이터 시트 생성 시작")

            # 새 시트 생성
            yearly_data_sheet_name = '연도별예산데이터'
            try:
                wb.sheets[yearly_data_sheet_name].delete()
                logging.info("기존 연도별예산데이터 시트 삭제")
            except:
                pass

            ws_yearly_data = wb.sheets.add(yearly_data_sheet_name)
            logging.info(f"새 연도별예산데이터 시트 생성: {yearly_data_sheet_name}")

            # 헤더 설정
            ws_yearly_data.range('A1').value = '연도'
            ws_yearly_data.range('B1').value = '예산과목'
            ws_yearly_data.range('C1').value = '예산금액'

            # 데이터 변환
            row_idx = 2  # 데이터 시작 행

            for year, budget_data in YEARLY_BUDGET_DATA.items():
                for budget_item, amount in budget_data.items():
                    ws_yearly_data.range(f'A{row_idx}').value = year
                    ws_yearly_data.range(f'B{row_idx}').value = budget_item
                    ws_yearly_data.range(f'C{row_idx}').value = amount
                    row_idx += 1

            logging.info(f"연도별 예산 세로형 데이터 생성 완료: {row_idx-1}개 행")

            # 즉시 저장하여 데이터 확정
            try:
                wb.save()
                logging.info("연도별 예산 데이터 시트 저장 완료")
            except Exception as save_error:
                logging.warning(f"연도별 예산 데이터 시트 저장 실패: {save_error}")

            return yearly_data_sheet_name

        except Exception as e:
            logging.error(f"연도별 예산 데이터 시트 생성 중 오류: {str(e)}")
            return None

    def _create_yearly_pivot_table_in_dashboard(self, wb, source_sheet_name: str, ws_dashboard) -> object:
        '''연도별 예산 비교용 피벗 테이블을 대시보드 시트의 B50에 생성합니다.'''
        try:
            import xlwings as xw

            # 연도별 데이터 시트를 소스로 사용
            ws_source = wb.sheets[source_sheet_name]
            
            # 데이터 범위 확인 (A:연도, B:예산과목, C:예산금액)
            try:
                used_range = ws_source.api.UsedRange
                last_row = used_range.Row + used_range.Rows.Count - 1
                source_range = ws_source.range(f'A1:C{last_row}')
                logging.info(f"연도별 데이터 범위: {source_range.address}")
            except:
                # 수동으로 범위 찾기
                last_row = 1
                for row in range(1, 1000):  # 최대 1000행까지 확인
                    if ws_source.range(f'A{row}').value is not None:
                        last_row = row
                source_range = ws_source.range(f'A1:C{last_row}')
                logging.info(f"수동 범위 설정: {source_range.address}")

            logging.info(f"소스 데이터 범위: {source_range.address}")

            # 1. 피벗 캐시 생성
            pivot_cache = wb.api.PivotCaches().Create(
                SourceType=xw.constants.PivotTableSourceType.xlDatabase,
                SourceData=source_range.api
            )

            # 2. 피벗 테이블을 대시보드 시트의 B53에 생성
            logging.info("대시보드 시트에 연도별 피벗 테이블 생성 중...")
            pivot_table = pivot_cache.CreatePivotTable(
                TableDestination=ws_dashboard.range('B53').api,
                TableName='YearlyBudgetComparisonDashboard'
            )
            logging.info("대시보드 시트에 연도별 피벗 테이블 기본 구조 생성 완료")

            # 3. 필드 배치
            logging.info("필드 배치 시작...")
            
            # 행 필드: 예산과목
            logging.info("예산과목 필드를 행 필드로 설정 중...")
            pivot_table.PivotFields('예산과목').Orientation = xw.constants.PivotFieldOrientation.xlRowField
            logging.info("예산과목 행 필드 설정 완료")

            # 열 필드: 연도
            logging.info("연도 필드를 열 필드로 설정 중...")
            pivot_table.PivotFields('연도').Orientation = xw.constants.PivotFieldOrientation.xlColumnField
            logging.info("연도 열 필드 설정 완료")

            # 값 필드: 예산금액
            logging.info("예산금액 필드를 데이터 필드로 설정 중...")
            try:
                # 사용 가능한 필드 확인
                available_fields = [field.Name for field in pivot_table.PivotFields()]
                logging.info(f"피벗 테이블 사용 가능한 필드들: {available_fields}")
                
                # 예산금액 필드 추가
                budget_field = pivot_table.PivotFields('예산금액')
                data_field = pivot_table.AddDataField(
                    budget_field,
                    '예산금액 합계',
                    xw.constants.ConsolidationFunction.xlSum
                )
                logging.info("예산금액 데이터 필드 추가 완료")
                
                # 총합계 행과 열 제거
                logging.info("총합계 행과 열 제거 중...")
                try:
                    # 행 총합계 제거
                    pivot_table.RowGrand = False
                    logging.info("행 총합계 제거 완료")
                    
                    # 열 총합계 제거
                    pivot_table.ColumnGrand = False
                    logging.info("열 총합계 제거 완료")
                except Exception as grand_error:
                    logging.warning(f"총합계 제거 실패: {grand_error}")
                
            except Exception as field_error:
                logging.error(f"필드 설정 중 오류: {field_error}")
                return None

            # 피벗 테이블 스타일 설정
            try:
                pivot_table.TableStyle2 = 'PivotStyleMedium9'
                logging.info("피벗 테이블 스타일 적용 완료")
            except Exception as style_error:
                logging.warning(f"피벗 테이블 스타일 적용 실패: {style_error}")

            logging.info("대시보드 시트에 연도별 예산 비교 피벗 테이블 생성 완료")
            return pivot_table

        except Exception as e:
            logging.error(f"대시보드 시트 연도별 피벗 테이블 생성 중 오류: {str(e)}")
            return None

    def _create_yearly_pivot_table(self, wb, source_sheet_name: str, ws_pivot) -> object:
        '''연도별 예산 비교용 피벗 테이블을 생성합니다.'''
        try:
            import xlwings as xw

            # 연도별 데이터 시트를 소스로 사용
            ws_source = wb.sheets[source_sheet_name]
            
            # 데이터 범위 확인 (A:연도, B:예산과목, C:예산금액)
            try:
                used_range = ws_source.api.UsedRange
                last_row = used_range.Row + used_range.Rows.Count - 1
                source_range = ws_source.range(f'A1:C{last_row}')
                logging.info(f"연도별 데이터 범위: {source_range.address}")
            except:
                # 수동으로 범위 찾기
                last_row = 1
                for row in range(1, 1000):  # 최대 1000행까지 확인
                    if ws_source.range(f'A{row}').value is not None:
                        last_row = row
                source_range = ws_source.range(f'A1:C{last_row}')
                logging.info(f"수동 범위 설정: {source_range.address}")

            logging.info(f"소스 데이터 범위: {source_range.address}")

            # 1. 피벗 캐시 생성
            pivot_cache = wb.api.PivotCaches().Create(
                SourceType=xw.constants.PivotTableSourceType.xlDatabase,
                SourceData=source_range.api
            )

            # 2. 피벗 테이블 생성
            logging.info("연도별 피벗 테이블 생성 중...")
            pivot_table = pivot_cache.CreatePivotTable(
                TableDestination=ws_pivot.range('B5').api,
                TableName='YearlyBudgetComparison'
            )
            logging.info("연도별 피벗 테이블 기본 구조 생성 완료")

            # 3. 필드 배치
            logging.info("필드 배치 시작...")
            
            # 행 필드: 예산과목
            logging.info("예산과목 필드를 행 필드로 설정 중...")
            pivot_table.PivotFields('예산과목').Orientation = xw.constants.PivotFieldOrientation.xlRowField
            logging.info("예산과목 행 필드 설정 완료")

            # 열 필드: 연도
            logging.info("연도 필드를 열 필드로 설정 중...")
            pivot_table.PivotFields('연도').Orientation = xw.constants.PivotFieldOrientation.xlColumnField
            logging.info("연도 열 필드 설정 완료")

            # 값 필드: 예산금액
            logging.info("예산금액 필드를 데이터 필드로 설정 중...")
            try:
                # 사용 가능한 필드 확인
                available_fields = [field.Name for field in pivot_table.PivotFields()]
                logging.info(f"피벗 테이블 사용 가능한 필드들: {available_fields}")
                
                # 예산금액 필드 추가
                budget_field = pivot_table.PivotFields('예산금액')
                data_field = pivot_table.AddDataField(
                    budget_field,
                    '예산금액 합계',
                    xw.constants.ConsolidationFunction.xlSum
                )
                logging.info("예산금액 데이터 필드 추가 완료")
                
                # 총합계 행과 열 제거
                logging.info("총합계 행과 열 제거 중...")
                try:
                    # 행 총합계 제거
                    pivot_table.RowGrand = False
                    logging.info("행 총합계 제거 완료")
                    
                    # 열 총합계 제거
                    pivot_table.ColumnGrand = False
                    logging.info("열 총합계 제거 완료")
                except Exception as grand_error:
                    logging.warning(f"총합계 제거 실패: {grand_error}")
                
            except Exception as field_error:
                logging.error(f"필드 설정 중 오류: {field_error}")
                return None

            # 피벗 테이블 스타일 설정
            try:
                pivot_table.TableStyle2 = 'PivotStyleMedium9'
                logging.info("피벗 테이블 스타일 적용 완료")
            except Exception as style_error:
                logging.warning(f"피벗 테이블 스타일 적용 실패: {style_error}")

            logging.info("연도별 예산 비교 피벗 테이블 생성 완료")
            return pivot_table

        except Exception as e:
            logging.error(f"연도별 피벗 테이블 생성 중 오류: {str(e)}")
            return None

    def _add_yearly_comparison_chart_to_dashboard(self, wb, ws_dashboard, pivot_table):
        '''대시보드 시트에 연도별 예산 비교 차트를 추가합니다.'''
        try:
            import xlwings as xw

            # 피벗 차트 생성 (세로 막대형)
            chart_shape = ws_dashboard.api.Shapes.AddChart2(
                227,  # 차트 스타일
                xw.constants.ChartType.xlColumnClustered
            )
            chart = chart_shape.Chart

            # 피벗 테이블 범위를 차트 소스로 설정
            # 연도별예산비교 시트의 피벗테이블(B5부터)을 소스로 사용
            source_range = wb.sheets['연도별예산비교'].range('B5').expand()
            chart.SetSourceData(Source=source_range.api)

            # 차트 제목 설정
            chart.HasTitle = True
            chart.ChartTitle.Text = '연도별 예산과목별 예산금액 비교'

            # 축 제목 설정
            try:
                chart.Axes(1).HasTitle = True  # X축
                chart.Axes(1).AxisTitle.Text = '예산과목'
                chart.Axes(2).HasTitle = True  # Y축
                chart.Axes(2).AxisTitle.Text = '예산금액 (원)'
            except Exception as axis_error:
                logging.warning(f"축 제목 설정 실패: {axis_error}")

            # 정확한 위치 지정: B53 셀의 좌상단 좌표에 맞춤 (포인트 단위)
            anchor_cell = ws_dashboard.range('B53')
            chart_shape.Left = anchor_cell.left
            chart_shape.Top = anchor_cell.top
            # 차트 크기 지정 (포인트)
            chart_shape.Width = 800
            chart_shape.Height = 400

            logging.info("대시보드 시트에 연도별 예산 비교 차트 추가 완료 (B53 정렬)")

        except Exception as e:
            logging.error(f"대시보드 시트 연도별 비교 차트 추가 중 오류: {str(e)}")

    def _add_yearly_comparison_chart(self, ws_pivot, pivot_table):
        '''연도별 예산 비교 차트를 추가합니다.'''
        try:
            import xlwings as xw

            # 피벗 차트 생성 (세로 막대형)
            chart_shape = ws_pivot.api.Shapes.AddChart2(
                227,  # 차트 스타일
                xw.constants.ChartType.xlColumnClustered
            )
            chart = chart_shape.Chart

            # 피벗 테이블 범위를 차트 소스로 설정 (이 함수는 피벗 시트 전용)
            source_range = ws_pivot.range('B5').expand()
            chart.SetSourceData(Source=source_range.api)

            # 차트 제목/축 설정
            chart.HasTitle = True
            chart.ChartTitle.Text = '연도별 예산과목별 예산금액 비교'
            try:
                chart.Axes(1).HasTitle = True
                chart.Axes(1).AxisTitle.Text = '예산과목'
                chart.Axes(2).HasTitle = True
                chart.Axes(2).AxisTitle.Text = '예산금액 (원)'
            except Exception as axis_error:
                logging.warning(f"축 제목 설정 실패: {axis_error}")

            # 기본 위치/크기
            chart_shape.Left = 500
            chart_shape.Top = 50
            chart_shape.Width = 500
            chart_shape.Height = 350

            logging.info("연도별 예산 비교 차트 추가 완료")

        except Exception as e:
            logging.error(f"연도별 비교 차트 추가 중 오류: {str(e)}")

    def _add_yearly_slicers_to_dashboard(self, wb, pivot_table, ws_dashboard):
        '''대시보드 시트에 연도별 비교용 슬라이서를 추가합니다. (차트 오른쪽)'''
        try:
            # 차트 기준 좌표/크기 계산 (B53에 차트 좌상단, 800x400 크기)
            anchor_cell = ws_dashboard.range('B53')
            chart_left = anchor_cell.left
            chart_top = anchor_cell.top
            chart_width = 800
            chart_height = 400
            chart_right = chart_left + chart_width
            chart_bottom = chart_top + chart_height
            margin = 2

            # 1) 연도 슬라이서: 차트 오른쪽 상단 시작
            try:
                slicer_cache_year = wb.api.SlicerCaches.Add2(
                    pivot_table,
                    '연도'
                )
                year_left = chart_right + margin
                year_top = chart_top
                year_width = 120
                year_height = 135
                slicer_cache_year.Slicers.Add(
                    SlicerDestination=ws_dashboard.api,
                    Name='DashboardYearSlicer',
                    Caption='연도 선택',
                    Top=year_top,
                    Left=year_left,
                    Width=year_width,
                    Height=year_height
                )
                logging.info("대시보드 시트에 연도 슬라이서 추가 완료")
            except Exception as e:
                logging.warning(f"대시보드 시트 연도 슬라이서 추가 실패: {str(e)}")

            # 2) 예산과목 슬라이서: 연도 슬라이서 바로 아래에서 시작하여 차트 하단까지
            try:
                slicer_cache_budget = wb.api.SlicerCaches.Add2(
                    pivot_table,
                    '예산과목'
                )
                budget_left = year_left
                # 연도 슬라이서 바로 아래부터 시작
                budget_top = year_top + year_height + margin
                budget_width = year_width
                # 피벗차트의 하단까지 차지하도록 높이 계산 (최소 높이 보장)
                budget_height = max(60, int(chart_bottom - budget_top))
                slicer_cache_budget.Slicers.Add(
                    SlicerDestination=ws_dashboard.api,
                    Name='DashboardBudgetItemSlicer',
                    Caption='예산과목 선택',
                    Top=budget_top,
                    Left=budget_left,
                    Width=budget_width,
                    Height=budget_height
                )
                logging.info("대시보드 시트에 예산과목 슬라이서 추가 완료")
            except Exception as e:
                logging.warning(f"대시보드 시트 예산과목 슬라이서 추가 실패: {str(e)}")

        except Exception as e:
            logging.error(f"대시보드 시트 연도별 슬라이서 추가 중 오류: {str(e)}")

    def _add_yearly_slicers(self, wb, pivot_table, ws_pivot):
        '''연도별 비교용 슬라이서를 추가합니다.'''
        try:
            import xlwings as xw

            # 예산과목 슬라이서 추가
            try:
                slicer_cache_budget = wb.api.SlicerCaches.Add2(
                    pivot_table,
                    '예산과목'
                )
                
                slicer_cache_budget.Slicers.Add(
                    SlicerDestination=ws_pivot.api,
                    Name='YearlyBudgetItemSlicer',
                    Caption='예산과목 선택',
                    Top=50,
                    Left=1050,
                    Width=200,
                    Height=400
                )
                logging.info("예산과목 슬라이서 추가 완료")
            except Exception as e:
                logging.warning(f"예산과목 슬라이서 추가 실패: {str(e)}")

            # 연도 슬라이서 추가
            try:
                slicer_cache_year = wb.api.SlicerCaches.Add2(
                    pivot_table,
                    '연도'
                )
                
                slicer_cache_year.Slicers.Add(
                    SlicerDestination=ws_pivot.api,
                    Name='YearSlicer',
                    Caption='연도 선택',
                    Top=470,
                    Left=1050,
                    Width=200,
                    Height=150
                )
                logging.info("연도 슬라이서 추가 완료")
            except Exception as e:
                logging.warning(f"연도 슬라이서 추가 실패: {str(e)}")

        except Exception as e:
            logging.error(f"연도별 슬라이서 추가 중 오류: {str(e)}")

    def reorder_all_sheets(self, file_path: str) -> bool:
        '''모든 시트를 원하는 순서로 재배치합니다.'''
        if not self.xlwings_available:
            logging.info("xlwings를 사용할 수 없어 시트 순서 조정을 건너뜁니다.")
            return False

        try:
            import xlwings as xw

            logging.info("전체 시트 순서 조정 시작")

            with xw.App(visible=True, add_book=False) as app:
                wb = app.books.open(file_path)

                # 원하는 시트 순서 정의
                desired_order = [
                    '대시보드',
                    '총액',
                    '사업비',
                    '연구비',
                    '집행관리(사업비)',
                    '집행관리(연구비)',
                    '연도별예산비교',
                    '예산분석',
                    '연도별예산데이터',
                    '예산분석데이터'
                ]

                # 현재 시트 목록 확인
                current_sheets = [sheet.name for sheet in wb.sheets]
                logging.info(f"현재 시트 순서: {current_sheets}")

                # 원하는 순서대로 시트 이동
                for target_index, sheet_name in enumerate(desired_order):
                    if sheet_name in current_sheets:
                        # 시트 찾기
                        sheet = wb.sheets[sheet_name]
                        
                        # 시트를 원하는 위치로 이동 (xlwings에서는 before 파라미터 사용)
                        if target_index == 0:
                            # 첫 번째 위치로 이동
                            sheet.api.Move(Before=wb.sheets[0].api)
                        else:
                            # 특정 위치 다음으로 이동
                            previous_sheet_name = desired_order[target_index - 1]
                            if previous_sheet_name in current_sheets:
                                previous_sheet = wb.sheets[previous_sheet_name]
                                sheet.api.Move(After=previous_sheet.api)
                        
                        logging.debug(f"시트 '{sheet_name}' 위치 조정 완료")
                    else:
                        logging.debug(f"시트 '{sheet_name}'를 찾을 수 없습니다.")

                # 최종 시트 순서 확인
                final_sheets = [sheet.name for sheet in wb.sheets]
                logging.info(f"최종 시트 순서: {final_sheets}")

                wb.save()
                logging.info("시트 순서 조정 완료")
                return True

        except Exception as e:
            logging.error(f"전체 시트 순서 조정 중 오류: {str(e)}")
            return False